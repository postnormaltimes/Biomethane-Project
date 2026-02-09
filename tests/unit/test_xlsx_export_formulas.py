from pathlib import Path

from openpyxl import load_workbook

from dcf_engine.engine import DCFEngine
from dcf_engine.models import (
    DCFInputs,
    TimelineInputs,
    RevenueInputs,
    OperatingInputs,
    NWCInputs,
    InvestmentInputs,
    TaxInputs,
    CAPMInputs,
    DebtInputs,
    WACCInputs,
    EquityBookInputs,
    TerminalValueInputs,
    NetDebtInputs,
    DiscountingMode,
    TerminalValueMethod,
    WeightingMode,
)
from dcf_io.writers import export_xlsx


def _make_golden_inputs() -> DCFInputs:
    return DCFInputs(
        discounting_mode=DiscountingMode.YEAR_SPECIFIC_FLAT,
        timeline=TimelineInputs(base_year=2022, forecast_years=[2023, 2024, 2025]),
        revenue=RevenueInputs(
            base_revenue=12500.0,
            growth_rates={2023: 0.15, 2024: 0.10, 2025: 0.10},
        ),
        operating=OperatingInputs(
            cost_ratios={2023: 0.85, 2024: 0.83, 2025: 0.80},
            depreciation_amortization={2022: 500.0, 2023: 550.0, 2024: 650.0, 2025: 700.0},
        ),
        nwc=NWCInputs(
            nwc_percent={2022: 0.16, 2023: 0.16, 2024: 0.13, 2025: 0.10},
        ),
        investments=InvestmentInputs(capex={2023: 800.0, 2024: 900.0, 2025: 1000.0}),
        tax=TaxInputs(tax_rate=0.30),
        capm=CAPMInputs(rf=0.04, rm=0.10, beta=1.30),
        debt=DebtInputs(
            debt_balances={2022: 1500.0, 2023: 2050.0, 2024: 2055.63, 2025: 2039.38},
            rd={2022: 0.05, 2023: 0.06, 2024: 0.065, 2025: 0.065},
        ),
        wacc=WACCInputs(
            weighting_mode=WeightingMode.BOOK_VALUE,
            equity_book_inputs=EquityBookInputs(base_equity_book=10000.0),
        ),
        terminal_value=TerminalValueInputs(method=TerminalValueMethod.PERPETUITY, g=0.0),
        net_debt=NetDebtInputs(cash_and_equivalents=1492.10),
    )


def test_export_xlsx_formulas(tmp_path: Path) -> None:
    inputs = _make_golden_inputs()
    outputs = DCFEngine(inputs).run()
    out_path = tmp_path / "dcf_formulas.xlsx"

    export_xlsx(outputs, out_path, xlsx_mode="formulas")

    wb = load_workbook(out_path, data_only=False)
    assert "Assumptions" in wb.sheetnames
    assert "Cash_Flow" in wb.sheetnames
    assert "Discounting" in wb.sheetnames
    assert "Valuation_Summary" in wb.sheetnames
    assert "Audit_Notes" in wb.sheetnames

    cash_flow = wb["Cash_Flow"]
    assert isinstance(cash_flow["B6"].value, str)
    assert cash_flow["B6"].value.startswith("=")

    discounting = wb["Discounting"]
    assert isinstance(discounting["B8"].value, str)
    assert discounting["B8"].value.startswith("=")

    assumptions = wb["Assumptions"]
    assert not (isinstance(assumptions["B2"].value, str) and assumptions["B2"].value.startswith("="))
