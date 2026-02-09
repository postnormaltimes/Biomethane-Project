from pathlib import Path

import yaml
from openpyxl import load_workbook

from dcf_projects.biometano.schema import BiometanoCase
from dcf_projects.biometano.builder import build_projections
from dcf_projects.biometano.statements import build_statements
from dcf_projects.biometano.valuation import compute_valuation
from dcf_io.writers import export_xlsx_biometano
from dcf_io.xlsx_validation import FormulaCheck, find_missing_formulas


CASE_FILE = Path(__file__).parent.parent.parent / "src/dcf_projects/biometano/case_files/biometano_case.yaml"


def test_biometano_export_formulas(tmp_path: Path) -> None:
    with open(CASE_FILE) as f:
        data = yaml.safe_load(f)
    case = BiometanoCase.model_validate(data)

    projections = build_projections(case)
    statements = build_statements(case, projections)
    valuation = compute_valuation(case, projections)

    out_path = tmp_path / "biometano_formulas.xlsx"
    export_xlsx_biometano(
        projections,
        statements,
        valuation,
        out_path,
        case=case,
        xlsx_mode="formulas",
    )

    wb = load_workbook(out_path, data_only=False)
    assert "Assumptions" in wb.sheetnames
    assert "Revenue_By_Channel" in wb.sheetnames
    assert "Discounting" in wb.sheetnames
    assert "Valuation_Summary" in wb.sheetnames
    assert "Audit_Notes" in wb.sheetnames
    assert "Audit_Checks" in wb.sheetnames
    assert "Balance_Sheet_Reclass" in wb.sheetnames

    missing = find_missing_formulas(
        wb,
        [
            FormulaCheck("Revenue_By_Channel", ["B2"]),
            FormulaCheck("Audit_Checks", ["B2"]),
            FormulaCheck("Balance_Sheet_Reclass", ["B4"]),
        ],
    )
    assert not missing

    assumptions = wb["Assumptions"]
    assert not (isinstance(assumptions["B2"].value, str) and assumptions["B2"].value.startswith("="))
