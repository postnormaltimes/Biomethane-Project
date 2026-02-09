"""Workbook validation helpers for formula-driven exports."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class FormulaCheck:
    sheet: str
    cells: Iterable[str]


def load_workbook_formulas(path: str):
    return load_workbook(path, data_only=False)


def find_missing_formulas(wb, checks: Iterable[FormulaCheck]) -> list[str]:
    missing: list[str] = []
    for check in checks:
        ws = wb[check.sheet]
        for cell in check.cells:
            value = ws[cell].value
            if not (isinstance(value, str) and value.startswith("=")):
                missing.append(f"{check.sheet}!{cell}")
    return missing


def find_non_formula_cells(wb, sheet: str, cells: Iterable[str]) -> list[str]:
    ws = wb[sheet]
    non_formula: list[str] = []
    for cell in cells:
        value = ws[cell].value
        if value is None:
            continue
        if isinstance(value, str) and value.startswith("="):
            continue
        non_formula.append(f"{sheet}!{cell}")
    return non_formula
