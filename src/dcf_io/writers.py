"""
DCF I/O Writers

Export to XLSX and CSV formats.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from dcf_engine.models import DCFOutputs
from dcf_projects.biometano.schema import BiometanoCase
from dcf_io.xlsx_layout import SheetLayout


def _create_inputs_summary(outputs: DCFOutputs) -> pd.DataFrame:
    """Create inputs/assumptions summary table."""
    data = [
        ["Base Year", outputs.base_year],
        ["Forecast Years", ", ".join(map(str, outputs.forecast_years))],
        ["Discounting Mode", outputs.discounting_mode.value],
        ["Cost of Equity (Ke)", f"{outputs.ke:.4f}"],
        ["Terminal Value Method", outputs.terminal_value.method.value],
    ]
    
    if outputs.terminal_value.growth_rate is not None:
        data.append(["Terminal Growth Rate", f"{outputs.terminal_value.growth_rate:.4f}"])
    
    return pd.DataFrame(data, columns=["Parameter", "Value"])


def _create_projections_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create operating projections table."""
    rows = []
    for p in outputs.projections:
        rows.append({
            "Year": p.year,
            "Revenue": p.revenue,
            "Operating Costs": p.operating_costs,
            "EBITDA": p.ebitda,
            "D&A": p.depreciation_amortization,
            "EBIT": p.ebit,
            "Tax on EBIT": p.tax_on_ebit,
            "NOPAT": p.nopat,
            "Interest Expense": p.interest_expense,
            "EBT": p.ebt,
            "Taxes on EBT": p.taxes_on_ebt,
            "Net Income": p.net_income,
        })
    return pd.DataFrame(rows)


def _create_nwc_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create NWC schedule table."""
    rows = []
    for n in outputs.nwc_schedule:
        rows.append({
            "Year": n.year,
            "NWC": n.nwc,
            "ΔNWC": n.delta_nwc,
        })
    return pd.DataFrame(rows)


def _create_investments_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create investment schedule table."""
    rows = []
    for cf in outputs.cash_flows:
        rows.append({
            "Year": cf.year,
            "Capex": cf.capex,
        })
    return pd.DataFrame(rows)


def _create_cashflows_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create cash flows table."""
    rows = []
    for cf in outputs.cash_flows:
        rows.append({
            "Year": cf.year,
            "NOPAT": cf.nopat,
            "D&A": cf.depreciation_amortization,
            "ΔNWC": cf.delta_nwc,
            "Capex": cf.capex,
            "FCFF": cf.fcff,
            "Interest Expense": cf.interest_expense,
            "Interest Tax Shield": cf.interest_tax_shield,
            "Net Borrowing": cf.net_borrowing,
            "FCFE": cf.fcfe,
        })
    return pd.DataFrame(rows)


def _create_discount_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create discount factors and PV table."""
    rows = []
    for d in outputs.discount_schedule:
        rows.append({
            "Year": d.year,
            "Period": d.period,
            "WACC": d.wacc,
            "Ke": d.ke,
            "DF (WACC)": d.discount_factor_wacc,
            "DF (Ke)": d.discount_factor_ke,
            "FCFF": d.fcff,
            "PV(FCFF)": d.pv_fcff,
            "FCFE": d.fcfe,
            "PV(FCFE)": d.pv_fcfe,
        })
    
    # Add totals row
    sum_pv_fcff = sum(d.pv_fcff for d in outputs.discount_schedule)
    sum_pv_fcfe = sum(d.pv_fcfe for d in outputs.discount_schedule)
    rows.append({
        "Year": "Total",
        "Period": "",
        "WACC": "",
        "Ke": "",
        "DF (WACC)": "",
        "DF (Ke)": "",
        "FCFF": "",
        "PV(FCFF)": sum_pv_fcff,
        "FCFE": "",
        "PV(FCFE)": sum_pv_fcfe,
    })
    
    return pd.DataFrame(rows)


def _create_terminal_value_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create terminal value computation table."""
    tv = outputs.terminal_value
    data = [
        ["Final Year", tv.final_year],
        ["Method", tv.method.value],
        ["Growth Rate (g)", tv.growth_rate if tv.growth_rate is not None else "N/A"],
        ["Exit Multiple", tv.exit_multiple if tv.exit_multiple is not None else "N/A"],
        ["Exit Metric", tv.exit_metric if tv.exit_metric else "N/A"],
        ["---", "---"],
        ["Terminal Value (FCFF)", tv.terminal_value_fcff],
        ["Terminal Value (FCFE)", tv.terminal_value_fcfe],
        ["Discount Rate (WACC)", tv.discount_rate_wacc],
        ["Discount Rate (Ke)", tv.discount_rate_ke],
        ["PV of TV (FCFF)", tv.pv_terminal_value_fcff],
        ["PV of TV (FCFE)", tv.pv_terminal_value_fcfe],
    ]
    return pd.DataFrame(data, columns=["Item", "Value"])


def _create_valuation_bridge_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create valuation bridge table."""
    vb = outputs.valuation_bridge
    data = [
        ["FCFF/WACC Approach", ""],
        ["Sum PV(FCFF)", vb.sum_pv_fcff],
        ["PV(Terminal Value)", vb.pv_terminal_value_fcff],
        ["Enterprise Value", vb.enterprise_value],
        ["Less: Debt at Base", -vb.debt_at_base],
        ["Plus: Cash at Base", vb.cash_at_base],
        ["Net Debt", vb.net_debt],
        ["Equity Value (from EV)", vb.equity_value_from_ev],
        ["---", "---"],
        ["FCFE/Ke Approach", ""],
        ["Sum PV(FCFE)", vb.sum_pv_fcfe],
        ["PV(Terminal Value)", vb.pv_terminal_value_fcfe],
        ["Equity Value (Direct)", vb.equity_value_direct],
        ["---", "---"],
        ["Reconciliation", ""],
        ["Difference", vb.reconciliation_difference],
    ]
    return pd.DataFrame(data, columns=["Item", "Value"])


def _create_reconciliation_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create FCFF vs FCFE reconciliation report."""
    vb = outputs.valuation_bridge
    
    data = [
        ["Equity from FCFF/WACC", vb.equity_value_from_ev],
        ["Equity from FCFE/Ke", vb.equity_value_direct],
        ["Difference", vb.reconciliation_difference],
        ["Difference (%)", abs(vb.reconciliation_difference / vb.equity_value_from_ev * 100) if vb.equity_value_from_ev else 0],
    ]
    
    # Add notes
    for i, note in enumerate(vb.reconciliation_notes):
        data.append([f"Note {i+1}", note])
    
    return pd.DataFrame(data, columns=["Item", "Value"])


def _create_wacc_details_table(outputs: DCFOutputs) -> pd.DataFrame:
    """Create WACC details table."""
    rows = []
    for w in outputs.wacc_details:
        rows.append({
            "Year": w.year,
            "Ke": w.ke,
            "Rd": w.rd,
            "Tax Rate": w.tax_rate,
            "Debt": w.debt,
            "Equity Book": w.equity_book,
            "Weight Debt": w.weight_debt,
            "Weight Equity": w.weight_equity,
            "WACC": w.wacc,
        })
    return pd.DataFrame(rows)


def format_tables(outputs: DCFOutputs) -> dict[str, pd.DataFrame]:
    """
    Convert DCF outputs to display-ready DataFrames.
    
    Returns:
        Dict mapping table name to DataFrame
    """
    return {
        "1_Inputs_Summary": _create_inputs_summary(outputs),
        "2_Operating_Projections": _create_projections_table(outputs),
        "3_NWC_Schedule": _create_nwc_table(outputs),
        "4_Investments": _create_investments_table(outputs),
        "5_Cash_Flows": _create_cashflows_table(outputs),
        "6_WACC_Details": _create_wacc_details_table(outputs),
        "7_PV_Decomposition": _create_discount_table(outputs),
        "8_Terminal_Value": _create_terminal_value_table(outputs),
        "9_Valuation_Bridge": _create_valuation_bridge_table(outputs),
        "10_Reconciliation": _create_reconciliation_table(outputs),
    }


def _style_xlsx_sheet(ws, df: pd.DataFrame):
    """Apply styling to Excel worksheet."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Style data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00' if isinstance(cell.value, float) else '#,##0'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width


def _style_table_header(ws, header_row: int, max_col: int) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")


def _style_table_body(ws, start_row: int, end_row: int, end_col: int) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, max_col=end_col):
        for cell in row:
            cell.border = thin_border


def _auto_fit_columns(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col, max_row=ws.max_row):
            cell = row[0]
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                continue
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)


def _apply_number_format(ws, rows: list[int], cols: list[int], fmt: str) -> None:
    for row in rows:
        for col in cols:
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = fmt


def _write_year_header(ws, years: list[int], layout: SheetLayout, title: str) -> None:
    ws.cell(row=layout.header_row, column=1, value=title)
    for year, col in layout.year_to_col(years).items():
        ws.cell(row=layout.header_row, column=col, value=year)


def _write_audit_notes(ws, formula_sheets: list[str], exceptions: list[str], conventions: list[str]) -> None:
    ws["A1"] = "Audit Notes"
    ws["A3"] = "Formula-driven sheets"
    for idx, name in enumerate(formula_sheets, start=4):
        ws.cell(row=idx, column=1, value=name)
    row = 4 + len(formula_sheets) + 1
    ws.cell(row=row, column=1, value="Value-only blocks/exceptions")
    if not exceptions:
        ws.cell(row=row + 1, column=1, value="None")
        row += 2
    else:
        for idx, note in enumerate(exceptions, start=row + 1):
            ws.cell(row=idx, column=1, value=note)
        row = row + 1 + len(exceptions)
    ws.cell(row=row + 1, column=1, value="Key conventions")
    for idx, note in enumerate(conventions, start=row + 2):
        ws.cell(row=idx, column=1, value=note)
    _auto_fit_columns(ws, 2)


def _write_dcf_formula_workbook(outputs: DCFOutputs, wb: Workbook) -> None:
    years = outputs.forecast_years
    all_years = [outputs.base_year] + years
    layout = SheetLayout()

    def _label(label: str, unit: str | None = None) -> str:
        return f"{label} ({unit})" if unit else label

    assumptions = wb.create_sheet("Assumptions")
    assumptions["A1"] = "Parameter"
    assumptions["B1"] = "Value"
    label_base_year = _label("Base Year", "year")
    label_ke = _label("Cost of Equity (Ke)", "%")
    label_g = _label("Terminal Growth Rate (g)", "%")
    label_exit_multiple = _label("Exit Multiple", "x")
    label_debt_base = _label("Debt at Base", "€")
    label_cash_base = _label("Cash at Base", "€")
    label_nwc_base = _label("NWC Base", "€")
    label_fixed_assets_base = _label("Fixed Assets Base", "€")

    params = [
        (label_base_year, outputs.base_year),
        ("Discounting Mode", outputs.discounting_mode.value),
        (label_ke, outputs.ke),
        ("Terminal Value Method", outputs.terminal_value.method.value),
        (label_g, outputs.terminal_value.growth_rate or 0.0),
        (label_exit_multiple, outputs.terminal_value.exit_multiple or 0.0),
        ("Exit Metric", outputs.terminal_value.exit_metric or ""),
        (label_debt_base, outputs.valuation_bridge.debt_at_base),
        (label_cash_base, outputs.valuation_bridge.cash_at_base),
        (label_nwc_base, outputs.nwc_schedule[0].nwc if outputs.nwc_schedule else 0.0),
        (label_fixed_assets_base, 0.0),
    assumptions = wb.create_sheet("Assumptions")
    assumptions["A1"] = "Parameter"
    assumptions["B1"] = "Value"
    params = [
        ("Base Year", outputs.base_year),
        ("Discounting Mode", outputs.discounting_mode.value),
        ("Cost of Equity (Ke)", outputs.ke),
        ("Terminal Value Method", outputs.terminal_value.method.value),
        ("Terminal Growth Rate (g)", outputs.terminal_value.growth_rate or 0.0),
        ("Exit Multiple", outputs.terminal_value.exit_multiple or 0.0),
        ("Exit Metric", outputs.terminal_value.exit_metric or ""),
        ("Debt at Base", outputs.valuation_bridge.debt_at_base),
        ("Cash at Base", outputs.valuation_bridge.cash_at_base),
    ]
    param_rows: dict[str, int] = {}
    for idx, (label, value) in enumerate(params, start=2):
        assumptions.cell(row=idx, column=1, value=label)
        assumptions.cell(row=idx, column=2, value=value)
        param_rows[label] = idx

    series_header_row = len(params) + 4
    assumptions.cell(row=series_header_row - 1, column=1, value="Series Inputs")
    series_layout = SheetLayout(start_col=2, header_row=series_header_row)
    assumptions.cell(row=series_header_row, column=1, value="Item")
    for year, col in series_layout.year_to_col(all_years).items():
        assumptions.cell(row=series_header_row, column=col, value=year)

    series_rows: dict[str, int] = {}

    def _series_row(label: str, values: dict[int, float]) -> None:
        row = series_header_row + 1 + len(series_rows)
        series_rows[label] = row
        assumptions.cell(row=row, column=1, value=label)
        for year, col in series_layout.year_to_col(all_years).items():
            if year in values:
                assumptions.cell(row=row, column=col, value=values[year])

    label_revenue = _label("Revenue", "€")
    label_operating_costs = _label("Operating Costs", "€")
    label_da = _label("D&A", "€")
    label_interest = _label("Interest Expense", "€")
    label_capex = _label("Capex", "€")
    label_net_borrowing = _label("Net Borrowing", "€")
    label_nwc_pct = _label("NWC % of Revenue", "%")
    label_tax_rate = _label("Tax Rate", "%")
    label_wacc = _label("WACC", "%")
    label_debt_balance = _label("Debt Balance", "€")
    label_equity_book = _label("Equity Book", "€")

    _series_row(label_revenue, {p.year: p.revenue for p in outputs.projections})
    _series_row(label_operating_costs, {p.year: p.operating_costs for p in outputs.projections})
    _series_row(label_da, {p.year: p.depreciation_amortization for p in outputs.projections})
    _series_row(label_interest, {p.year: p.interest_expense for p in outputs.projections})
    _series_row(label_capex, {c.year: c.capex for c in outputs.cash_flows})
    _series_row(label_net_borrowing, {c.year: c.net_borrowing for c in outputs.cash_flows})
    _series_row(
        label_nwc_pct,
        {
            p.year: (outputs.nwc_schedule[idx + 1].nwc / p.revenue) if p.revenue else 0.0
            for idx, p in enumerate(outputs.projections)
        },
    )
    _series_row(label_tax_rate, {w.year: w.tax_rate for w in outputs.wacc_details})
    _series_row(label_wacc, {d.year: d.wacc for d in outputs.discount_schedule})
    _series_row(label_debt_balance, {w.year: w.debt for w in outputs.wacc_details})
    equity_values = outputs.equity_book_values or {}
    _series_row(label_equity_book, equity_values)
    _series_row("Revenue", {p.year: p.revenue for p in outputs.projections})
    _series_row("Operating Costs", {p.year: p.operating_costs for p in outputs.projections})
    _series_row("D&A", {p.year: p.depreciation_amortization for p in outputs.projections})
    _series_row("Interest Expense", {p.year: p.interest_expense for p in outputs.projections})
    _series_row("Capex", {c.year: c.capex for c in outputs.cash_flows})
    _series_row("Net Borrowing", {c.year: c.net_borrowing for c in outputs.cash_flows})
    _series_row("NWC", {n.year: n.nwc for n in outputs.nwc_schedule})
    _series_row("Tax Rate", {w.year: w.tax_rate for w in outputs.wacc_details})
    _series_row("WACC", {d.year: d.wacc for d in outputs.discount_schedule})
    _series_row("Debt Balance", {w.year: w.debt for w in outputs.wacc_details})
    equity_values = outputs.equity_book_values or {}
    _series_row("Equity Book", equity_values)

    _style_table_header(assumptions, 1, 2)
    _style_table_header(assumptions, series_header_row, 1 + len(all_years))
    _style_table_body(assumptions, 2, assumptions.max_row, 1 + len(all_years))
    _auto_fit_columns(assumptions, 1 + len(all_years))
    _apply_number_format(assumptions, [4, 5, 6, 8, 9], [2], "0.0000")

    def _param_cell(label: str) -> str:
        return f"B{param_rows[label]}"

    def _assumption_cell(label: str, year: int) -> str:
        row = series_rows[label]
        col = series_layout.year_to_col(all_years)[year]
        return series_layout.cell(col, row)

    revenue_sheet = wb.create_sheet("Revenue_By_Channel")
    _write_year_header(revenue_sheet, years, layout, "Line Item")
    revenue_sheet.cell(row=2, column=1, value=_label("Total Revenue", "€"))
    for year, col in layout.year_to_col(years).items():
        revenue_sheet.cell(row=2, column=col, value=f"=Assumptions!{_assumption_cell(label_revenue, year)}")
    revenue_sheet.cell(row=2, column=1, value="Total Revenue")
    for year, col in layout.year_to_col(years).items():
        revenue_sheet.cell(row=2, column=col, value=f"=Assumptions!{_assumption_cell('Revenue', year)}")
    _style_table_header(revenue_sheet, 1, 1 + len(years))
    _style_table_body(revenue_sheet, 2, 2, 1 + len(years))
    _auto_fit_columns(revenue_sheet, 1 + len(years))

    opex_sheet = wb.create_sheet("OPEX")
    _write_year_header(opex_sheet, years, layout, "Line Item")
    opex_sheet.cell(row=2, column=1, value=_label("Operating Costs", "€"))
    opex_sheet.cell(row=3, column=1, value=_label("EBITDA", "€"))
    for year, col in layout.year_to_col(years).items():
        opex_sheet.cell(row=2, column=col, value=f"=Assumptions!{_assumption_cell(label_operating_costs, year)}")
    opex_sheet.cell(row=2, column=1, value="Operating Costs")
    opex_sheet.cell(row=3, column=1, value="EBITDA")
    for year, col in layout.year_to_col(years).items():
        opex_sheet.cell(row=2, column=col, value=f"=Assumptions!{_assumption_cell('Operating Costs', year)}")
        rev_cell = layout.cell(col, 2)
        opex_sheet.cell(row=3, column=col, value=f"=Revenue_By_Channel!{rev_cell}-OPEX!{layout.cell(col, 2)}")
    _style_table_header(opex_sheet, 1, 1 + len(years))
    _style_table_body(opex_sheet, 2, 3, 1 + len(years))
    _auto_fit_columns(opex_sheet, 1 + len(years))

    income_sheet = wb.create_sheet("Income_Statement")
    _write_year_header(income_sheet, years, layout, "Line Item")
    labels = [
        _label("Revenue", "€"),
        _label("Operating Costs", "€"),
        _label("EBITDA", "€"),
        _label("D&A", "€"),
        _label("EBIT", "€"),
        _label("Tax on EBIT", "€"),
        _label("NOPAT", "€"),
        _label("Interest Expense", "€"),
        _label("EBT", "€"),
        _label("Taxes on EBT", "€"),
        _label("Net Income", "€"),
        "Revenue",
        "Operating Costs",
        "EBITDA",
        "D&A",
        "EBIT",
        "Tax on EBIT",
        "NOPAT",
        "Interest Expense",
        "EBT",
        "Taxes on EBT",
        "Net Income",
    ]
    for idx, label in enumerate(labels, start=2):
        income_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        income_sheet.cell(row=2, column=col, value=f"=Revenue_By_Channel!{layout.cell(col, 2)}")
        income_sheet.cell(row=3, column=col, value=f"=OPEX!{layout.cell(col, 2)}")
        income_sheet.cell(row=4, column=col, value=f"=OPEX!{layout.cell(col, 3)}")
        income_sheet.cell(row=5, column=col, value=f"=Assumptions!{_assumption_cell(label_da, year)}")
        income_sheet.cell(row=6, column=col, value=f"=Income_Statement!{layout.cell(col, 4)}-Income_Statement!{layout.cell(col, 5)}")
        tax_rate = f"Assumptions!{_assumption_cell(label_tax_rate, year)}"
        income_sheet.cell(row=7, column=col, value=f"=Income_Statement!{layout.cell(col, 6)}*{tax_rate}")
        income_sheet.cell(row=8, column=col, value=f"=Income_Statement!{layout.cell(col, 6)}-Income_Statement!{layout.cell(col, 7)}")
        income_sheet.cell(row=9, column=col, value=f"=Assumptions!{_assumption_cell(label_interest, year)}")
        income_sheet.cell(row=5, column=col, value=f"=Assumptions!{_assumption_cell('D&A', year)}")
        income_sheet.cell(row=6, column=col, value=f"=Income_Statement!{layout.cell(col, 4)}-Income_Statement!{layout.cell(col, 5)}")
        tax_rate = f"Assumptions!{_assumption_cell('Tax Rate', year)}"
        income_sheet.cell(row=7, column=col, value=f"=Income_Statement!{layout.cell(col, 6)}*{tax_rate}")
        income_sheet.cell(row=8, column=col, value=f"=Income_Statement!{layout.cell(col, 6)}-Income_Statement!{layout.cell(col, 7)}")
        income_sheet.cell(row=9, column=col, value=f"=Assumptions!{_assumption_cell('Interest Expense', year)}")
        income_sheet.cell(row=10, column=col, value=f"=Income_Statement!{layout.cell(col, 6)}-Income_Statement!{layout.cell(col, 9)}")
        income_sheet.cell(row=11, column=col, value=f"=Income_Statement!{layout.cell(col, 10)}*{tax_rate}")
        income_sheet.cell(row=12, column=col, value=f"=Income_Statement!{layout.cell(col, 10)}-Income_Statement!{layout.cell(col, 11)}")
    _style_table_header(income_sheet, 1, 1 + len(years))
    _style_table_body(income_sheet, 2, 12, 1 + len(years))
    _auto_fit_columns(income_sheet, 1 + len(years))

    balance_sheet = wb.create_sheet("Balance_Sheet")
    balance_layout = SheetLayout(start_col=2, header_row=1)
    _write_year_header(balance_sheet, all_years, balance_layout, "Line Item")
    balance_sheet.cell(row=2, column=1, value=_label("NWC", "€"))
    balance_sheet.cell(row=3, column=1, value=_label("ΔNWC", "€"))
    balance_sheet.cell(row=4, column=1, value=_label("Debt Balance", "€"))
    balance_sheet.cell(row=5, column=1, value=_label("Equity Book", "€"))
    for year, col in balance_layout.year_to_col(all_years).items():
        if year == outputs.base_year:
            balance_sheet.cell(row=2, column=col, value=f"=Assumptions!{_param_cell(label_nwc_base)}")
        else:
            balance_sheet.cell(
                row=2,
                column=col,
                value=f"=Revenue_By_Channel!{layout.cell(col, 2)}*Assumptions!{_assumption_cell(label_nwc_pct, year)}",
            )
    balance_sheet.cell(row=2, column=1, value="NWC")
    balance_sheet.cell(row=3, column=1, value="ΔNWC")
    balance_sheet.cell(row=4, column=1, value="Debt Balance")
    balance_sheet.cell(row=5, column=1, value="Equity Book")
    for year, col in balance_layout.year_to_col(all_years).items():
        balance_sheet.cell(row=2, column=col, value=f"=Assumptions!{_assumption_cell('NWC', year)}")
        if year == outputs.base_year:
            balance_sheet.cell(row=3, column=col, value="")
        else:
            prev_year = all_years[all_years.index(year) - 1]
            prev_col = balance_layout.year_to_col(all_years)[prev_year]
            balance_sheet.cell(row=3, column=col, value=f"=Balance_Sheet!{balance_layout.cell(col, 2)}-Balance_Sheet!{balance_layout.cell(prev_col, 2)}")
        if year in years:
            balance_sheet.cell(row=4, column=col, value=f"=Assumptions!{_assumption_cell(label_debt_balance, year)}")
        if year in equity_values:
            balance_sheet.cell(row=5, column=col, value=f"=Assumptions!{_assumption_cell(label_equity_book, year)}")
            balance_sheet.cell(row=4, column=col, value=f"=Assumptions!{_assumption_cell('Debt Balance', year)}")
        if year in equity_values:
            balance_sheet.cell(row=5, column=col, value=f"=Assumptions!{_assumption_cell('Equity Book', year)}")
    _style_table_header(balance_sheet, 1, 1 + len(all_years))
    _style_table_body(balance_sheet, 2, 5, 1 + len(all_years))
    _auto_fit_columns(balance_sheet, 1 + len(all_years))

    reclass_sheet = wb.create_sheet("Balance_Sheet_Reclass")
    _write_year_header(reclass_sheet, all_years, balance_layout, "Line Item")
    reclass_labels = [
        _label("Fixed Assets (Net)", "€"),
        _label("NWC", "€"),
        _label("CIN", "€"),
        _label("Debt", "€"),
        _label("Cash", "€"),
        _label("NFP", "€"),
        _label("Equity", "€"),
        _label("CIN - (Equity + NFP)", "€"),
    ]
    for idx, label in enumerate(reclass_labels, start=2):
        reclass_sheet.cell(row=idx, column=1, value=label)
    for year, col in balance_layout.year_to_col(all_years).items():
        if year == outputs.base_year:
            reclass_sheet.cell(row=2, column=col, value=f"=Assumptions!{_param_cell(label_fixed_assets_base)}")
        else:
            prev_year = all_years[all_years.index(year) - 1]
            prev_col = balance_layout.year_to_col(all_years)[prev_year]
            reclass_sheet.cell(
                row=2,
                column=col,
                value=(
                    f"=Balance_Sheet_Reclass!{balance_layout.cell(prev_col, 2)}"
                    f"+Assumptions!{_assumption_cell(label_capex, year)}"
                    f"-Assumptions!{_assumption_cell(label_da, year)}"
                ),
            )
        reclass_sheet.cell(row=3, column=col, value=f"=Balance_Sheet!{balance_layout.cell(col, 2)}")
        reclass_sheet.cell(row=4, column=col, value=f"=Balance_Sheet_Reclass!{balance_layout.cell(col, 2)}+Balance_Sheet_Reclass!{balance_layout.cell(col, 3)}")
        if year == outputs.base_year:
            reclass_sheet.cell(row=5, column=col, value=f"=Balance_Sheet!{balance_layout.cell(col, 4)}")
            reclass_sheet.cell(row=6, column=col, value=f"=Assumptions!{_param_cell(label_cash_base)}")
        else:
            reclass_sheet.cell(row=5, column=col, value=f"=Balance_Sheet!{balance_layout.cell(col, 4)}")
            prev_year = all_years[all_years.index(year) - 1]
            prev_col = balance_layout.year_to_col(all_years)[prev_year]
            reclass_sheet.cell(row=6, column=col, value=f"=Balance_Sheet_Reclass!{balance_layout.cell(prev_col, 6)}")
        reclass_sheet.cell(row=7, column=col, value=f"=Balance_Sheet_Reclass!{balance_layout.cell(col, 5)}-Balance_Sheet_Reclass!{balance_layout.cell(col, 6)}")
        reclass_sheet.cell(row=8, column=col, value=f"=Balance_Sheet!{balance_layout.cell(col, 5)}")
        reclass_sheet.cell(
            row=9,
            column=col,
            value=f"=Balance_Sheet_Reclass!{balance_layout.cell(col, 4)}-(Balance_Sheet_Reclass!{balance_layout.cell(col, 7)}+Balance_Sheet_Reclass!{balance_layout.cell(col, 8)})",
        )
    _style_table_header(reclass_sheet, 1, 1 + len(all_years))
    _style_table_body(reclass_sheet, 2, 9, 1 + len(all_years))
    _auto_fit_columns(reclass_sheet, 1 + len(all_years))

    cash_flow = wb.create_sheet("Cash_Flow")
    _write_year_header(cash_flow, years, layout, "Line Item")
    cf_labels = [
        _label("NOPAT", "€"),
        _label("D&A", "€"),
        _label("ΔNWC", "€"),
        _label("Capex", "€"),
        _label("FCFF", "€"),
        _label("Interest Expense", "€"),
        _label("Interest Tax Shield", "€"),
        _label("Net Borrowing", "€"),
        _label("FCFE", "€"),
    cash_flow = wb.create_sheet("Cash_Flow")
    _write_year_header(cash_flow, years, layout, "Line Item")
    cf_labels = [
        "NOPAT",
        "D&A",
        "ΔNWC",
        "Capex",
        "FCFF",
        "Interest Expense",
        "Interest Tax Shield",
        "Net Borrowing",
        "FCFE",
    ]
    for idx, label in enumerate(cf_labels, start=2):
        cash_flow.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        tax_rate = f"Assumptions!{_assumption_cell(label_tax_rate, year)}"
        cash_flow.cell(row=2, column=col, value=f"=Income_Statement!{layout.cell(col, 8)}")
        cash_flow.cell(row=3, column=col, value=f"=Assumptions!{_assumption_cell(label_da, year)}")
        cash_flow.cell(row=4, column=col, value=f"=Balance_Sheet!{balance_layout.cell(balance_layout.year_to_col(all_years)[year], 3)}")
        cash_flow.cell(row=5, column=col, value=f"=Assumptions!{_assumption_cell(label_capex, year)}")
        cash_flow.cell(row=6, column=col, value=f"=Cash_Flow!{layout.cell(col, 2)}+Cash_Flow!{layout.cell(col, 3)}-Cash_Flow!{layout.cell(col, 4)}-Cash_Flow!{layout.cell(col, 5)}")
        cash_flow.cell(row=7, column=col, value=f"=Income_Statement!{layout.cell(col, 9)}")
        cash_flow.cell(row=8, column=col, value=f"=Cash_Flow!{layout.cell(col, 7)}*{tax_rate}")
        cash_flow.cell(row=9, column=col, value=f"=Assumptions!{_assumption_cell(label_net_borrowing, year)}")
        tax_rate = f"Assumptions!{_assumption_cell('Tax Rate', year)}"
        cash_flow.cell(row=2, column=col, value=f"=Income_Statement!{layout.cell(col, 8)}")
        cash_flow.cell(row=3, column=col, value=f"=Assumptions!{_assumption_cell('D&A', year)}")
        cash_flow.cell(row=4, column=col, value=f"=Balance_Sheet!{balance_layout.cell(balance_layout.year_to_col(all_years)[year], 3)}")
        cash_flow.cell(row=5, column=col, value=f"=Assumptions!{_assumption_cell('Capex', year)}")
        cash_flow.cell(row=6, column=col, value=f"=Cash_Flow!{layout.cell(col, 2)}+Cash_Flow!{layout.cell(col, 3)}-Cash_Flow!{layout.cell(col, 4)}-Cash_Flow!{layout.cell(col, 5)}")
        cash_flow.cell(row=7, column=col, value=f"=Income_Statement!{layout.cell(col, 9)}")
        cash_flow.cell(row=8, column=col, value=f"=Cash_Flow!{layout.cell(col, 7)}*{tax_rate}")
        cash_flow.cell(row=9, column=col, value=f"=Assumptions!{_assumption_cell('Net Borrowing', year)}")
        cash_flow.cell(
            row=10,
            column=col,
            value=f"=Cash_Flow!{layout.cell(col, 6)}-Cash_Flow!{layout.cell(col, 7)}+Cash_Flow!{layout.cell(col, 8)}+Cash_Flow!{layout.cell(col, 9)}",
        )
    _style_table_header(cash_flow, 1, 1 + len(years))
    _style_table_body(cash_flow, 2, 10, 1 + len(years))
    _auto_fit_columns(cash_flow, 1 + len(years))

    fcff_sheet = wb.create_sheet("FCFF")
    _write_year_header(fcff_sheet, years, layout, "Line Item")
    fcff_sheet.cell(row=2, column=1, value=_label("FCFF", "€"))
    fcff_sheet.cell(row=3, column=1, value=_label("FCFE", "€"))
    fcff_sheet.cell(row=2, column=1, value="FCFF")
    fcff_sheet.cell(row=3, column=1, value="FCFE")
    for year, col in layout.year_to_col(years).items():
        fcff_sheet.cell(row=2, column=col, value=f"=Cash_Flow!{layout.cell(col, 6)}")
        fcff_sheet.cell(row=3, column=col, value=f"=Cash_Flow!{layout.cell(col, 10)}")
    _style_table_header(fcff_sheet, 1, 1 + len(years))
    _style_table_body(fcff_sheet, 2, 3, 1 + len(years))
    _auto_fit_columns(fcff_sheet, 1 + len(years))

    discounting = wb.create_sheet("Discounting")
    _write_year_header(discounting, years, layout, "Line Item")
    disc_labels = [
        _label("Period", "year"),
        _label("WACC", "%"),
        _label("Ke", "%"),
        _label("DF (WACC)", "x"),
        _label("DF (Ke)", "x"),
        _label("FCFF", "€"),
        _label("PV(FCFF)", "€"),
        _label("FCFE", "€"),
        _label("PV(FCFE)", "€"),
        "Period",
        "WACC",
        "Ke",
        "DF (WACC)",
        "DF (Ke)",
        "FCFF",
        "PV(FCFF)",
        "FCFE",
        "PV(FCFE)",
    ]
    for idx, label in enumerate(disc_labels, start=2):
        discounting.cell(row=idx, column=1, value=label)
    for idx, (year, col) in enumerate(layout.year_to_col(years).items(), start=1):
        discounting.cell(row=2, column=col, value=idx)
        discounting.cell(row=3, column=col, value=f"=Assumptions!{_assumption_cell(label_wacc, year)}")
        discounting.cell(row=4, column=col, value=f"=Assumptions!{_param_cell(label_ke)}")
        discounting.cell(row=3, column=col, value=f"=Assumptions!{_assumption_cell('WACC', year)}")
        discounting.cell(row=4, column=col, value=f"=Assumptions!{_param_cell('Cost of Equity (Ke)')}")
        discounting.cell(row=5, column=col, value=f"=1/(1+Discounting!{layout.cell(col, 3)})^Discounting!{layout.cell(col, 2)}")
        discounting.cell(row=6, column=col, value=f"=1/(1+Discounting!{layout.cell(col, 4)})^Discounting!{layout.cell(col, 2)}")
        discounting.cell(row=7, column=col, value=f"=Cash_Flow!{layout.cell(col, 6)}")
        discounting.cell(row=8, column=col, value=f"=Discounting!{layout.cell(col, 7)}*Discounting!{layout.cell(col, 5)}")
        discounting.cell(row=9, column=col, value=f"=Cash_Flow!{layout.cell(col, 10)}")
        discounting.cell(row=10, column=col, value=f"=Discounting!{layout.cell(col, 9)}*Discounting!{layout.cell(col, 6)}")
    _style_table_header(discounting, 1, 1 + len(years))
    _style_table_body(discounting, 2, 10, 1 + len(years))
    _auto_fit_columns(discounting, 1 + len(years))

    valuation_sheet = wb.create_sheet("Valuation_Summary")
    valuation_sheet["A1"] = "Metric"
    valuation_sheet["B1"] = "Value"
    valuation_items = [
        _label("Sum PV(FCFF)", "€"),
        _label("Terminal Value (FCFF)", "€"),
        _label("PV Terminal Value (FCFF)", "€"),
        _label("Enterprise Value", "€"),
        _label("Debt at Base", "€"),
        _label("Cash at Base", "€"),
        _label("Net Debt", "€"),
        _label("Equity Value (from EV)", "€"),
        _label("Sum PV(FCFE)", "€"),
        _label("Terminal Value (FCFE)", "€"),
        _label("PV Terminal Value (FCFE)", "€"),
        _label("Equity Value (Direct)", "€"),
        _label("Reconciliation Difference", "€"),
        "Sum PV(FCFF)",
        "Terminal Value (FCFF)",
        "PV Terminal Value (FCFF)",
        "Enterprise Value",
        "Debt at Base",
        "Cash at Base",
        "Net Debt",
        "Equity Value (from EV)",
        "Sum PV(FCFE)",
        "Terminal Value (FCFE)",
        "PV Terminal Value (FCFE)",
        "Equity Value (Direct)",
        "Reconciliation Difference",
    ]
    for idx, label in enumerate(valuation_items, start=2):
        valuation_sheet.cell(row=idx, column=1, value=label)
    last_col = layout.year_to_col(years)[years[-1]]
    tv_method = outputs.terminal_value.method.value
    if tv_method == "perpetuity":
        tv_formula = (
            f"=Cash_Flow!{layout.cell(last_col, 6)}*(1+Assumptions!{_param_cell(label_g)})"
            f"/(Assumptions!{_assumption_cell(label_wacc, years[-1])}-Assumptions!{_param_cell(label_g)})"
        )
        tv_fcfe_formula = (
            f"=Cash_Flow!{layout.cell(last_col, 10)}*(1+Assumptions!{_param_cell(label_g)})"
            f"/(Assumptions!{_param_cell(label_ke)}-Assumptions!{_param_cell(label_g)})"
            f"=Cash_Flow!{layout.cell(last_col, 6)}*(1+Assumptions!{_param_cell('Terminal Growth Rate (g)')})"
            f"/(Assumptions!{_assumption_cell('WACC', years[-1])}-Assumptions!{_param_cell('Terminal Growth Rate (g)')})"
        )
        tv_fcfe_formula = (
            f"=Cash_Flow!{layout.cell(last_col, 10)}*(1+Assumptions!{_param_cell('Terminal Growth Rate (g)')})"
            f"/(Assumptions!{_param_cell('Cost of Equity (Ke)')}-Assumptions!{_param_cell('Terminal Growth Rate (g)')})"
        )
    else:
        metric = (outputs.terminal_value.exit_metric or "EBITDA").lower()
        if metric == "revenue":
            metric_cell = f"Revenue_By_Channel!{layout.cell(last_col, 2)}"
        elif metric == "ebit":
            metric_cell = f"Income_Statement!{layout.cell(last_col, 6)}"
        else:
            metric_cell = f"OPEX!{layout.cell(last_col, 3)}"
        tv_formula = f"={metric_cell}*Assumptions!{_param_cell(label_exit_multiple)}"
        tv_formula = f"={metric_cell}*Assumptions!{_param_cell('Exit Multiple')}"
        tv_fcfe_formula = tv_formula

    valuation_sheet.cell(
        row=2,
        column=2,
        value=f"=SUM(Discounting!{layout.cell(layout.start_col, 8)}:{layout.cell(last_col, 8)})",
    )
    valuation_sheet.cell(row=3, column=2, value=tv_formula)
    valuation_sheet.cell(
        row=4,
        column=2,
        value=f"=Valuation_Summary!B3*Discounting!{layout.cell(last_col, 5)}",
    )
    valuation_sheet.cell(row=5, column=2, value="=Valuation_Summary!B2+Valuation_Summary!B4")
    valuation_sheet.cell(row=6, column=2, value=f"=Assumptions!{_param_cell(label_debt_base)}")
    valuation_sheet.cell(row=7, column=2, value=f"=Assumptions!{_param_cell(label_cash_base)}")
    valuation_sheet.cell(row=6, column=2, value=f"=Assumptions!{_param_cell('Debt at Base')}")
    valuation_sheet.cell(row=7, column=2, value=f"=Assumptions!{_param_cell('Cash at Base')}")
    valuation_sheet.cell(row=8, column=2, value="=Valuation_Summary!B6-Valuation_Summary!B7")
    valuation_sheet.cell(row=9, column=2, value="=Valuation_Summary!B5-Valuation_Summary!B8")
    valuation_sheet.cell(
        row=10,
        column=2,
        value=f"=SUM(Discounting!{layout.cell(layout.start_col, 10)}:{layout.cell(last_col, 10)})",
    )
    valuation_sheet.cell(row=11, column=2, value=tv_fcfe_formula)
    valuation_sheet.cell(
        row=12,
        column=2,
        value=f"=Valuation_Summary!B11*Discounting!{layout.cell(last_col, 6)}",
    )
    valuation_sheet.cell(row=13, column=2, value="=Valuation_Summary!B10+Valuation_Summary!B12")
    valuation_sheet.cell(row=14, column=2, value="=Valuation_Summary!B9-Valuation_Summary!B13")
    _style_table_header(valuation_sheet, 1, 2)
    _style_table_body(valuation_sheet, 2, 14, 2)
    _auto_fit_columns(valuation_sheet, 2)

    audit_checks = wb.create_sheet("Audit_Checks")
    _write_year_header(audit_checks, all_years, balance_layout, "Check")
    audit_checks.cell(row=2, column=1, value=_label("FCFF Identity", "€"))
    audit_checks.cell(row=3, column=1, value=_label("CIN Identity", "€"))
    audit_checks.cell(row=4, column=1, value=_label("PV Roll-up", "€"))
    for year, col in balance_layout.year_to_col(all_years).items():
        if year == outputs.base_year:
            audit_checks.cell(row=2, column=col, value="")
        else:
            audit_checks.cell(
                row=2,
                column=col,
                value=(
                    f"=Cash_Flow!{layout.cell(col, 6)}"
                    f"-(Income_Statement!{layout.cell(col, 6)}*(1-Assumptions!{_assumption_cell(label_tax_rate, year)})"
                    f"+Assumptions!{_assumption_cell(label_da, year)}"
                    f"-Assumptions!{_assumption_cell(label_capex, year)}"
                    f"-Balance_Sheet!{balance_layout.cell(col, 3)})"
                ),
            )
        audit_checks.cell(row=3, column=col, value=f"=Balance_Sheet_Reclass!{balance_layout.cell(col, 9)}")
        if year == years[0]:
            audit_checks.cell(row=4, column=col, value="=Valuation_Summary!B5-(Valuation_Summary!B2+Valuation_Summary!B4)")
        else:
            audit_checks.cell(row=4, column=col, value="")
    _style_table_header(audit_checks, 1, 1 + len(all_years))
    _style_table_body(audit_checks, 2, 4, 1 + len(all_years))
    _auto_fit_columns(audit_checks, 1 + len(all_years))

    audit = wb.create_sheet("Audit_Notes")
    _write_audit_notes(
        audit,
        [
            "Assumptions",
            "Revenue_By_Channel",
            "OPEX",
            "Income_Statement",
            "Balance_Sheet",
            "Balance_Sheet_Reclass",
            "Cash_Flow",
            "FCFF",
            "Discounting",
            "Valuation_Summary",
            "Audit_Checks",
        ],
        [
            "Assumptions series rows are inputs copied from model outputs (Revenue, OPEX, D&A, Capex, Interest, Net Borrowing, WACC, Tax Rate).",
            "NWC base and Fixed Assets base are input anchors for roll-forward schedules.",
            "Cash is held constant outside a modeled cash schedule.",
        ],
        [
            "Assumptions series rows are inputs copied from model outputs (Revenue, OPEX, D&A, Capex, Interest, Net Borrowing, NWC, WACC, Tax Rate).",
        ],
        [
            "Discount factors use end-of-period convention: DF=1/(1+r)^period.",
            "Terminal value uses perpetuity or exit multiple based on Assumptions.",
            "CIN check uses Balance_Sheet_Reclass: CIN = NWC + Fixed Assets; NFP = Debt - Cash.",
        ],
    )


def export_xlsx(outputs: DCFOutputs, path: str | Path, xlsx_mode: str = "formulas") -> None:
    """
    Export DCF outputs to Excel file.

    Args:
        outputs: DCF outputs to export
        path: Output file path
        xlsx_mode: "formulas" (default) or "values"
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_mode not in {"formulas", "values"}:
        raise ValueError("xlsx_mode must be 'formulas' or 'values'")

    if xlsx_mode == "values":
        tables = format_tables(outputs)
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, df in tables.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            _style_xlsx_sheet(ws, df)
        wb.save(path)
        return

    wb = Workbook()
    wb.remove(wb.active)
    _write_dcf_formula_workbook(outputs, wb)
    wb.save(path)


def export_csv(outputs: DCFOutputs, output_dir: str | Path) -> list[Path]:
    """
    Export DCF outputs to CSV files (one per table).
    
    Args:
        outputs: DCF outputs to export
        output_dir: Directory to write CSV files
    
    Returns:
        List of created file paths
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    tables = format_tables(outputs)
    created_files = []
    
    for table_name, df in tables.items():
        file_path = output_dir / f"{table_name}.csv"
        df.to_csv(file_path, index=False)
        created_files.append(file_path)
    
    return created_files


# ============================================================================
# BIOMETANO EXPORTS
# ============================================================================

def _create_biometano_production_table(projections) -> pd.DataFrame:
    """Create production table for biometano."""
    rows = []
    for p in projections.production:
        rows.append({
            "Year": p.year,
            "Availability": p.availability,
            "FORSU (t)": p.forsu_tonnes,
            "Biomethane (MWh)": p.biomethane_mwh,
            "CO2 (t)": p.co2_tonnes,
            "Compost (t)": p.compost_tonnes,
        })
    return pd.DataFrame(rows)


def _create_biometano_revenue_table(projections) -> pd.DataFrame:
    """Create revenue by channel table."""
    rows = []
    for r in projections.revenues:
        rows.append({
            "Year": r.year,
            "Gate Fee": r.gate_fee,
            "Tariff": r.tariff,
            "GO": r.go,
            "CO2": r.co2,
            "Compost": r.compost,
            "Total": r.total,
        })
    return pd.DataFrame(rows)


def _create_biometano_opex_table(projections) -> pd.DataFrame:
    """Create OPEX by category table."""
    rows = []
    for o in projections.opex:
        rows.append({
            "Year": o.year,
            "Feedstock": o.feedstock_handling,
            "Utilities": o.utilities,
            "Chemicals": o.chemicals,
            "Maintenance": o.maintenance,
            "Personnel": o.personnel,
            "Insurance": o.insurance,
            "Overheads": o.overheads,
            "Digestate": o.digestate_handling,
            "Other": o.other,
            "Total": o.total,
        })
    return pd.DataFrame(rows)


def _create_biometano_income_statement(statements) -> pd.DataFrame:
    """Create income statement table."""
    rows = []
    for s in statements.income_statements:
        rows.append({
            "Year": s.year,
            "Revenue": s.total_revenue,
            "OPEX": s.total_opex,
            "EBITDA": s.ebitda,
            "D&A": s.depreciation,
            "Grant Release": s.grant_income_release,
            "EBIT": s.ebit,
            "Interest": s.interest_expense,
            "EBT": s.ebt,
            "Taxes Before Credit": s.taxes_before_credit,
            "Tax Credit Used": s.tax_credit_utilization,
            "Taxes Paid": s.taxes_paid,
            "Net Income": s.net_income,
        })
    return pd.DataFrame(rows)


def _create_biometano_cash_flow(statements) -> pd.DataFrame:
    """Create cash flow statement table."""
    rows = []
    for c in statements.cash_flows:
        rows.append({
            "Year": c.year,
            "NOPAT": c.nopat,
            "D&A": c.depreciation,
            "Change in NWC": c.change_in_nwc,
            "CFO": c.cfo,
            "CAPEX": c.capex,
            "Grant Cash": c.grant_cash_received,
            "CFI": c.cfi,
            "Debt Draw": c.debt_drawdown,
            "Debt Repay": c.debt_repayment,
            "Interest Paid": c.interest_paid,
            "CFF": c.cff,
            "Net CF": c.net_cash_flow,
            "Closing Cash": c.closing_cash,
        })
    return pd.DataFrame(rows)


def _create_biometano_fcff_table(projections) -> pd.DataFrame:
    """Create FCFF breakdown table."""
    rows = []
    for year in projections.operating_years:
        rows.append({
            "Year": year,
            "EBIT": projections.ebit.get(year, 0),
            "D&A": projections.depreciation.get(year, 0),
            "ΔNWC": projections.delta_nwc.get(year, 0),
            "FCFF": projections.fcff.get(year, 0),
            "FCFE": projections.fcfe.get(year, 0),
        })
    return pd.DataFrame(rows)


def _create_biometano_valuation_table(valuation) -> pd.DataFrame:
    """Create valuation summary table."""
    data = [
        ["Cost of Equity (Ke)", valuation.ke],
        ["Sum PV(FCFF)", valuation.sum_pv_fcff],
        ["Terminal Value (FCFF)", valuation.terminal_value_fcff],
        ["PV Terminal Value", valuation.pv_terminal_value_fcff],
        ["Enterprise Value", valuation.enterprise_value],
        ["Debt at Base", valuation.debt_at_base],
        ["Cash at Base", valuation.cash_at_base],
        ["Net Debt", valuation.net_debt],
        ["Equity Value (FCFF/WACC)", valuation.equity_value],
        ["Equity Value (FCFE/Ke)", valuation.equity_value_direct],
        ["Reconciliation Difference", valuation.reconciliation_difference],
    ]
    return pd.DataFrame(data, columns=["Item", "Value"])


def format_biometano_tables(projections, statements, valuation) -> dict[str, pd.DataFrame]:
    """Format all biometano tables for export."""
    return {
        "1_Production": _create_biometano_production_table(projections),
        "2_Revenue": _create_biometano_revenue_table(projections),
        "3_OPEX": _create_biometano_opex_table(projections),
        "4_Income_Statement": _create_biometano_income_statement(statements),
        "5_Cash_Flow": _create_biometano_cash_flow(statements),
        "6_FCFF_FCFE": _create_biometano_fcff_table(projections),
        "7_Valuation": _create_biometano_valuation_table(valuation),
    }

def _write_biometano_formula_workbook(
    case: BiometanoCase,
    projections,
    statements,
    valuation,
    wb: Workbook,
) -> None:
    years = projections.all_forecast_years
    operating_years = projections.operating_years or years
    layout = SheetLayout()

    def _label(label: str, unit: str | None = None) -> str:
        return f"{label} ({unit})" if unit else label

    assumptions = wb.create_sheet("Assumptions")
    assumptions["A1"] = "Parameter"
    assumptions["B1"] = "Value"

    label_cod_year = _label("COD Year", "year")
    label_tax_rate = _label("Tax Rate", "%")
    label_ke = _label("Cost of Equity (Ke)", "%")
    label_g = _label("Terminal Growth Rate (g)", "%")
    label_exit_multiple = _label("Exit Multiple", "x")
    label_debt_base = _label("Debt at Base", "€")
    label_cash_base = _label("Cash at Base", "€")
    label_forsu = _label("FORSU Throughput", "t/y")
    label_biomethane = _label("Biomethane", "MWh/y")
    label_co2 = _label("CO2", "t/y")
    label_compost = _label("Compost", "t/y")
    label_gate_price = _label("Gate Fee Price", "€/t")
    label_gate_escal = _label("Gate Fee Escalation", "%")
    label_tariff_price = _label("Tariff Price", "€/MWh")
    label_tariff_escal = _label("Tariff Escalation", "%")
    label_go_price = _label("GO Price", "€/MWh")
    label_go_escal = _label("GO Escalation", "%")
    label_co2_price = _label("CO2 Price", "€/t")
    label_co2_escal = _label("CO2 Escalation", "%")
    label_compost_price = _label("Compost Price", "€/t")
    label_compost_escal = _label("Compost Escalation", "%")
    label_gate_dso = _label("Gate Fee DSO", "days")
    label_tariff_dso = _label("Tariff DSO", "days")
    label_go_dso = _label("GO DSO", "days")
    label_co2_dso = _label("CO2 DSO", "days")
    label_compost_dso = _label("Compost DSO", "days")
    label_feedstock_dpo = _label("Feedstock DPO", "days")
    label_utilities_dpo = _label("Utilities DPO", "days")
    label_chemicals_dpo = _label("Chemicals DPO", "days")
    label_maintenance_dpo = _label("Maintenance DPO", "days")
    label_personnel_dpo = _label("Personnel DPO", "days")
    label_insurance_dpo = _label("Insurance DPO", "days")
    label_overheads_dpo = _label("Overheads DPO", "days")
    label_digestate_dpo = _label("Digestate DPO", "days")
    label_other_dpo = _label("Other DPO", "days")

    params = [
        (_label("Base Year", "year"), case.horizon.base_year),
        (label_cod_year, case.horizon.cod_year),
        (label_tax_rate, case.financing.tax_rate),
        (label_ke, valuation.ke),
        ("Terminal Value Method", case.terminal_value.method),
        (label_g, case.terminal_value.perpetuity_growth),
        (label_exit_multiple, case.terminal_value.exit_multiple or 0.0),
        (label_debt_base, valuation.debt_at_base),
        (label_cash_base, valuation.cash_at_base),
        (label_forsu, case.production.forsu_throughput_tpy),
        (label_biomethane, case.production.get_biomethane_mwh()),
        (label_co2, case.production.co2_tpy),
        (label_compost, case.production.compost_tpy),
        (label_gate_price, case.revenues.gate_fee.price),
        (label_gate_escal, case.revenues.gate_fee.escalation_rate),
        (label_tariff_price, case.revenues.tariff.price),
        (label_tariff_escal, case.revenues.tariff.escalation_rate),
        (label_go_price, case.revenues.go.price),
        (label_go_escal, case.revenues.go.escalation_rate),
        (label_co2_price, case.revenues.co2.price),
        (label_co2_escal, case.revenues.co2.escalation_rate),
        (label_compost_price, case.revenues.compost.price),
        (label_compost_escal, case.revenues.compost.escalation_rate),
        (label_gate_dso, case.revenues.gate_fee.payment_delay_days),
        (label_tariff_dso, case.revenues.tariff.payment_delay_days),
        (label_go_dso, case.revenues.go.payment_delay_days),
        (label_co2_dso, case.revenues.co2.payment_delay_days),
        (label_compost_dso, case.revenues.compost.payment_delay_days),
        (label_feedstock_dpo, case.opex.feedstock_handling.payment_delay_days),
        (label_utilities_dpo, case.opex.utilities.payment_delay_days),
        (label_chemicals_dpo, case.opex.chemicals.payment_delay_days),
        (label_maintenance_dpo, case.opex.maintenance.payment_delay_days),
        (label_personnel_dpo, case.opex.personnel.payment_delay_days),
        (label_insurance_dpo, case.opex.insurance.payment_delay_days),
        (label_overheads_dpo, case.opex.overheads.payment_delay_days),
        (label_digestate_dpo, case.opex.digestate_handling.payment_delay_days),
        (label_other_dpo, case.opex.other.payment_delay_days),
    params = [
        ("Base Year", case.horizon.base_year),
        ("COD Year", case.horizon.cod_year),
        ("Tax Rate", case.financing.tax_rate),
        ("Cost of Equity (Ke)", valuation.ke),
        ("Terminal Value Method", case.terminal_value.method),
        ("Terminal Growth Rate (g)", case.terminal_value.perpetuity_growth),
        ("Exit Multiple", case.terminal_value.exit_multiple or 0.0),
        ("Debt at Base", valuation.debt_at_base),
        ("Cash at Base", valuation.cash_at_base),
        ("FORSU Throughput (tpy)", case.production.forsu_throughput_tpy),
        ("Biomethane (MWh/y)", case.production.get_biomethane_mwh()),
        ("CO2 (t/y)", case.production.co2_tpy),
        ("Compost (t/y)", case.production.compost_tpy),
        ("Gate Fee Price", case.revenues.gate_fee.price),
        ("Gate Fee Escalation", case.revenues.gate_fee.escalation_rate),
        ("Tariff Price", case.revenues.tariff.price),
        ("Tariff Escalation", case.revenues.tariff.escalation_rate),
        ("CO2 Price", case.revenues.co2.price),
        ("CO2 Escalation", case.revenues.co2.escalation_rate),
        ("GO Price", case.revenues.go.price),
        ("GO Escalation", case.revenues.go.escalation_rate),
        ("Compost Price", case.revenues.compost.price),
        ("Compost Escalation", case.revenues.compost.escalation_rate),
    ]
    param_rows: dict[str, int] = {}
    for idx, (label, value) in enumerate(params, start=2):
        assumptions.cell(row=idx, column=1, value=label)
        assumptions.cell(row=idx, column=2, value=value)
        param_rows[label] = idx

    series_header_row = len(params) + 4
    assumptions.cell(row=series_header_row - 1, column=1, value="Series Inputs")
    series_layout = SheetLayout(start_col=2, header_row=series_header_row)
    assumptions.cell(row=series_header_row, column=1, value="Item")
    for year, col in series_layout.year_to_col(years).items():
        assumptions.cell(row=series_header_row, column=col, value=year)

    series_rows: dict[str, int] = {}

    def _series_row(label: str, values: dict[int, float]) -> None:
        row = series_header_row + 1 + len(series_rows)
        series_rows[label] = row
        assumptions.cell(row=row, column=1, value=label)
        for year, col in series_layout.year_to_col(years).items():
            if year in values:
                assumptions.cell(row=row, column=col, value=values[year])

    label_availability = _label("Availability", "%")
    label_opex_feedstock = _label("OPEX Feedstock", "€")
    label_opex_utilities = _label("OPEX Utilities", "€")
    label_opex_chemicals = _label("OPEX Chemicals", "€")
    label_opex_maintenance = _label("OPEX Maintenance", "€")
    label_opex_personnel = _label("OPEX Personnel", "€")
    label_opex_insurance = _label("OPEX Insurance", "€")
    label_opex_overheads = _label("OPEX Overheads", "€")
    label_opex_digestate = _label("OPEX Digestate", "€")
    label_opex_other = _label("OPEX Other", "€")
    label_depreciation = _label("Depreciation", "€")
    label_grant_release = _label("Grant Release", "€")
    label_interest = _label("Interest Expense", "€")
    label_tax_credit = _label("Tax Credit Used", "€")
    label_delta_nwc = _label("ΔNWC", "€")
    label_capex = _label("Capex", "€")
    label_grant_cash = _label("Grant Cash", "€")
    label_debt_draw = _label("Debt Draw", "€")
    label_debt_repay = _label("Debt Repay", "€")
    label_interest_paid = _label("Interest Paid", "€")
    label_equity_contrib = _label("Equity Contribution", "€")
    label_opening_cash = _label("Opening Cash", "€")

    _series_row(label_availability, {p.year: p.availability for p in projections.production})
    _series_row(label_opex_feedstock, {o.year: o.feedstock_handling for o in projections.opex})
    _series_row(label_opex_utilities, {o.year: o.utilities for o in projections.opex})
    _series_row(label_opex_chemicals, {o.year: o.chemicals for o in projections.opex})
    _series_row(label_opex_maintenance, {o.year: o.maintenance for o in projections.opex})
    _series_row(label_opex_personnel, {o.year: o.personnel for o in projections.opex})
    _series_row(label_opex_insurance, {o.year: o.insurance for o in projections.opex})
    _series_row(label_opex_overheads, {o.year: o.overheads for o in projections.opex})
    _series_row(label_opex_digestate, {o.year: o.digestate_handling for o in projections.opex})
    _series_row(label_opex_other, {o.year: o.other for o in projections.opex})
    _series_row(label_depreciation, projections.depreciation)
    _series_row(label_grant_release, projections.grant_income_release)
    _series_row(label_interest, projections.interest)
    _series_row(label_tax_credit, projections.tax_credit_utilization)
    _series_row(label_delta_nwc, projections.delta_nwc)
    _series_row(label_capex, {c.year: -c.total for c in projections.capex})
    _series_row(label_grant_cash, {c.year: c.grant_cash_received for c in statements.cash_flows})
    _series_row(label_debt_draw, {c.year: c.debt_drawdown for c in statements.cash_flows})
    _series_row(label_debt_repay, {c.year: c.debt_repayment for c in statements.cash_flows})
    _series_row(label_interest_paid, {c.year: c.interest_paid for c in statements.cash_flows})
    _series_row(label_equity_contrib, {c.year: c.equity_contribution for c in statements.cash_flows})
    _series_row(label_opening_cash, {c.year: c.opening_cash for c in statements.cash_flows})

    label_cash = _label("Cash", "€")
    label_trade_receivables = _label("Trade Receivables", "€")
    label_grant_receivable = _label("Grant Receivable", "€")
    label_tax_credit_receivable = _label("Tax Credit Receivable", "€")
    label_fixed_assets_gross = _label("Fixed Assets Gross", "€")
    label_accum_depr = _label("Accumulated Depreciation", "€")
    label_trade_payables = _label("Trade Payables", "€")
    label_taxes_payable = _label("Taxes Payable", "€")
    label_debt = _label("Debt", "€")
    label_deferred_income = _label("Deferred Income", "€")
    label_share_capital = _label("Share Capital", "€")
    label_retained_earnings = _label("Retained Earnings", "€")
    label_current_profit = _label("Current Year Profit", "€")

    balance = {b.year: b for b in statements.balance_sheets}
    _series_row(label_cash, {y: b.cash for y, b in balance.items()})
    _series_row(label_trade_receivables, {y: b.trade_receivables for y, b in balance.items()})
    _series_row(label_grant_receivable, {y: b.grant_receivable for y, b in balance.items()})
    _series_row(label_tax_credit_receivable, {y: b.tax_credit_receivable for y, b in balance.items()})
    _series_row(label_fixed_assets_gross, {y: b.fixed_assets_gross for y, b in balance.items()})
    _series_row(label_accum_depr, {y: b.accumulated_depreciation for y, b in balance.items()})
    _series_row(label_trade_payables, {y: b.trade_payables for y, b in balance.items()})
    _series_row(label_taxes_payable, {y: b.taxes_payable for y, b in balance.items()})
    _series_row(label_debt, {y: b.debt for y, b in balance.items()})
    _series_row(label_deferred_income, {y: b.deferred_income for y, b in balance.items()})
    _series_row(label_share_capital, {y: b.share_capital for y, b in balance.items()})
    _series_row(label_retained_earnings, {y: b.retained_earnings for y, b in balance.items()})
    _series_row(label_current_profit, {y: b.current_year_profit for y, b in balance.items()})
    _series_row("Availability", {p.year: p.availability for p in projections.production})
    _series_row("OPEX Feedstock", {o.year: o.feedstock_handling for o in projections.opex})
    _series_row("OPEX Utilities", {o.year: o.utilities for o in projections.opex})
    _series_row("OPEX Chemicals", {o.year: o.chemicals for o in projections.opex})
    _series_row("OPEX Maintenance", {o.year: o.maintenance for o in projections.opex})
    _series_row("OPEX Personnel", {o.year: o.personnel for o in projections.opex})
    _series_row("OPEX Insurance", {o.year: o.insurance for o in projections.opex})
    _series_row("OPEX Overheads", {o.year: o.overheads for o in projections.opex})
    _series_row("OPEX Digestate", {o.year: o.digestate_handling for o in projections.opex})
    _series_row("OPEX Other", {o.year: o.other for o in projections.opex})
    _series_row("Depreciation", projections.depreciation)
    _series_row("Grant Release", projections.grant_income_release)
    _series_row("Interest Expense", projections.interest)
    _series_row("Tax Credit Used", projections.tax_credit_utilization)
    _series_row("Delta NWC", projections.delta_nwc)
    _series_row("Capex", {c.year: -c.total for c in projections.capex})
    _series_row("Grant Cash", {c.year: c.grant_cash_received for c in statements.cash_flows})
    _series_row("Debt Draw", {c.year: c.debt_drawdown for c in statements.cash_flows})
    _series_row("Debt Repay", {c.year: c.debt_repayment for c in statements.cash_flows})
    _series_row("Interest Paid", {c.year: c.interest_paid for c in statements.cash_flows})
    _series_row("Equity Contribution", {c.year: c.equity_contribution for c in statements.cash_flows})
    _series_row("Opening Cash", {c.year: c.opening_cash for c in statements.cash_flows})

    balance = {b.year: b for b in statements.balance_sheets}
    _series_row("Cash", {y: b.cash for y, b in balance.items()})
    _series_row("Trade Receivables", {y: b.trade_receivables for y, b in balance.items()})
    _series_row("Grant Receivable", {y: b.grant_receivable for y, b in balance.items()})
    _series_row("Tax Credit Receivable", {y: b.tax_credit_receivable for y, b in balance.items()})
    _series_row("Fixed Assets Gross", {y: b.fixed_assets_gross for y, b in balance.items()})
    _series_row("Accumulated Depreciation", {y: b.accumulated_depreciation for y, b in balance.items()})
    _series_row("Trade Payables", {y: b.trade_payables for y, b in balance.items()})
    _series_row("Taxes Payable", {y: b.taxes_payable for y, b in balance.items()})
    _series_row("Debt", {y: b.debt for y, b in balance.items()})
    _series_row("Deferred Income", {y: b.deferred_income for y, b in balance.items()})
    _series_row("Share Capital", {y: b.share_capital for y, b in balance.items()})
    _series_row("Retained Earnings", {y: b.retained_earnings for y, b in balance.items()})
    _series_row("Current Year Profit", {y: b.current_year_profit for y, b in balance.items()})

    _style_table_header(assumptions, 1, 2)
    _style_table_header(assumptions, series_header_row, 1 + len(years))
    _style_table_body(assumptions, 2, assumptions.max_row, 1 + len(years))
    _auto_fit_columns(assumptions, 1 + len(years))

    def _param_cell(label: str) -> str:
        return f"B{param_rows[label]}"

    def _series_cell(label: str, year: int) -> str:
        row = series_rows[label]
        col = series_layout.year_to_col(years)[year]
        return series_layout.cell(col, row)

    production_sheet = wb.create_sheet("Production")
    _write_year_header(production_sheet, years, layout, "Line Item")
    prod_labels = [
        _label("Availability", "%"),
        _label("FORSU", "t"),
        _label("Biomethane", "MWh"),
        _label("CO2", "t"),
        _label("Compost", "t"),
    ]
    for idx, label in enumerate(prod_labels, start=2):
        production_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        production_sheet.cell(row=2, column=col, value=f"=Assumptions!{_series_cell(label_availability, year)}")
    production_sheet.cell(row=3, column=col, value=f"=Assumptions!{_param_cell(label_forsu)}*Production!{layout.cell(col, 2)}")
    production_sheet.cell(row=4, column=col, value=f"=Assumptions!{_param_cell(label_biomethane)}*Production!{layout.cell(col, 2)}")
    production_sheet.cell(row=5, column=col, value=f"=Assumptions!{_param_cell(label_co2)}*Production!{layout.cell(col, 2)}")
    production_sheet.cell(row=6, column=col, value=f"=Assumptions!{_param_cell(label_compost)}*Production!{layout.cell(col, 2)}")
    prod_labels = ["Availability", "FORSU (t)", "Biomethane (MWh)", "CO2 (t)", "Compost (t)"]
    for idx, label in enumerate(prod_labels, start=2):
        production_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        production_sheet.cell(row=2, column=col, value=f"=Assumptions!{_series_cell('Availability', year)}")
        production_sheet.cell(row=3, column=col, value=f"=Assumptions!{_param_cell('FORSU Throughput (tpy)')}*Production!{layout.cell(col, 2)}")
        production_sheet.cell(row=4, column=col, value=f"=Assumptions!{_param_cell('Biomethane (MWh/y)')}*Production!{layout.cell(col, 2)}")
        production_sheet.cell(row=5, column=col, value=f"=Assumptions!{_param_cell('CO2 (t/y)')}*Production!{layout.cell(col, 2)}")
        production_sheet.cell(row=6, column=col, value=f"=Assumptions!{_param_cell('Compost (t/y)')}*Production!{layout.cell(col, 2)}")
    _style_table_header(production_sheet, 1, 1 + len(years))
    _style_table_body(production_sheet, 2, 6, 1 + len(years))
    _auto_fit_columns(production_sheet, 1 + len(years))

    revenue_sheet = wb.create_sheet("Revenue_By_Channel")
    _write_year_header(revenue_sheet, years, layout, "Line Item")
    rev_labels = [
        _label("Gate Fee", "€"),
        _label("Tariff", "€"),
        _label("GO", "€"),
        _label("CO2", "€"),
        _label("Compost", "€"),
        _label("Total", "€"),
    ]
    for idx, label in enumerate(rev_labels, start=2):
        revenue_sheet.cell(row=idx, column=1, value=label)
    cod_cell = _param_cell(label_cod_year)
    rev_labels = ["Gate Fee", "Tariff", "CO2", "GO", "Compost", "Total"]
    for idx, label in enumerate(rev_labels, start=2):
        revenue_sheet.cell(row=idx, column=1, value=label)
    cod_cell = _param_cell("COD Year")
    for year, col in layout.year_to_col(years).items():
        year_ref = layout.cell(col, 1)
        revenue_sheet.cell(
            row=2,
            column=col,
            value=(
                f"=IF({year_ref}<Assumptions!{cod_cell},0,"
                f"Production!{layout.cell(col, 3)}*Assumptions!{_param_cell(label_gate_price)}"
                f"*(1+Assumptions!{_param_cell(label_gate_escal)})^({year_ref}-Assumptions!{cod_cell}))"
                f"Production!{layout.cell(col, 3)}*Assumptions!{_param_cell('Gate Fee Price')}"
                f"*(1+Assumptions!{_param_cell('Gate Fee Escalation')})^({year_ref}-Assumptions!{cod_cell}))"
            ),
        )
        revenue_sheet.cell(
            row=3,
            column=col,
            value=(
                f"=IF({year_ref}<Assumptions!{cod_cell},0,"
                f"Production!{layout.cell(col, 4)}*Assumptions!{_param_cell(label_tariff_price)}"
                f"*(1+Assumptions!{_param_cell(label_tariff_escal)})^({year_ref}-Assumptions!{cod_cell}))"
                f"Production!{layout.cell(col, 4)}*Assumptions!{_param_cell('Tariff Price')}"
                f"*(1+Assumptions!{_param_cell('Tariff Escalation')})^({year_ref}-Assumptions!{cod_cell}))"
            ),
        )
        revenue_sheet.cell(
            row=4,
            column=col,
            value=(
                f"=IF({year_ref}<Assumptions!{cod_cell},0,"
                f"Production!{layout.cell(col, 4)}*Assumptions!{_param_cell(label_go_price)}"
                f"*(1+Assumptions!{_param_cell(label_go_escal)})^({year_ref}-Assumptions!{cod_cell}))"
                f"Production!{layout.cell(col, 5)}*Assumptions!{_param_cell('CO2 Price')}"
                f"*(1+Assumptions!{_param_cell('CO2 Escalation')})^({year_ref}-Assumptions!{cod_cell}))"
            ),
        )
        revenue_sheet.cell(
            row=5,
            column=col,
            value=(
                f"=IF({year_ref}<Assumptions!{cod_cell},0,"
                f"Production!{layout.cell(col, 5)}*Assumptions!{_param_cell(label_co2_price)}"
                f"*(1+Assumptions!{_param_cell(label_co2_escal)})^({year_ref}-Assumptions!{cod_cell}))"
                f"Production!{layout.cell(col, 4)}*Assumptions!{_param_cell('GO Price')}"
                f"*(1+Assumptions!{_param_cell('GO Escalation')})^({year_ref}-Assumptions!{cod_cell}))"
            ),
        )
        revenue_sheet.cell(
            row=6,
            column=col,
            value=(
                f"=IF({year_ref}<Assumptions!{cod_cell},0,"
                f"Production!{layout.cell(col, 6)}*Assumptions!{_param_cell(label_compost_price)}"
                f"*(1+Assumptions!{_param_cell(label_compost_escal)})^({year_ref}-Assumptions!{cod_cell}))"
                f"Production!{layout.cell(col, 6)}*Assumptions!{_param_cell('Compost Price')}"
                f"*(1+Assumptions!{_param_cell('Compost Escalation')})^({year_ref}-Assumptions!{cod_cell}))"
            ),
        )
        revenue_sheet.cell(
            row=7,
            column=col,
            value=f"=SUM(Revenue_By_Channel!{layout.cell(col, 2)}:{layout.cell(col, 6)})",
        )
    _style_table_header(revenue_sheet, 1, 1 + len(years))
    _style_table_body(revenue_sheet, 2, 7, 1 + len(years))
    _auto_fit_columns(revenue_sheet, 1 + len(years))

    opex_sheet = wb.create_sheet("OPEX")
    _write_year_header(opex_sheet, years, layout, "Line Item")
    opex_labels = [
        (_label("Feedstock", "€"), label_opex_feedstock),
        (_label("Utilities", "€"), label_opex_utilities),
        (_label("Chemicals", "€"), label_opex_chemicals),
        (_label("Maintenance", "€"), label_opex_maintenance),
        (_label("Personnel", "€"), label_opex_personnel),
        (_label("Insurance", "€"), label_opex_insurance),
        (_label("Overheads", "€"), label_opex_overheads),
        (_label("Digestate", "€"), label_opex_digestate),
        (_label("Other", "€"), label_opex_other),
        (_label("Total", "€"), None),
    ]
    for idx, (label, _) in enumerate(opex_labels, start=2):
        opex_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        for idx, (_, series_label) in enumerate(opex_labels, start=2):
            if series_label:
                opex_sheet.cell(row=idx, column=col, value=f"=Assumptions!{_series_cell(series_label, year)}")
        opex_sheet.cell(
            row=11,
            column=col,
            value=f"=SUM(OPEX!{layout.cell(col, 2)}:{layout.cell(col, 10)})",
        )
    _style_table_header(opex_sheet, 1, 1 + len(years))
    _style_table_body(opex_sheet, 2, 11, 1 + len(years))
    _auto_fit_columns(opex_sheet, 1 + len(years))

    income_sheet = wb.create_sheet("Income_Statement")
    _write_year_header(income_sheet, years, layout, "Line Item")
    income_labels = [
        _label("Revenue", "€"),
        _label("OPEX", "€"),
        _label("EBITDA", "€"),
        _label("D&A", "€"),
        _label("Grant Release", "€"),
        _label("EBIT", "€"),
        _label("Interest", "€"),
        _label("EBT", "€"),
        _label("Taxes Before Credit", "€"),
        _label("Tax Credit Used", "€"),
        _label("Taxes Paid", "€"),
        _label("Net Income", "€"),
    ]
    for idx, label in enumerate(income_labels, start=2):
        income_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        income_sheet.cell(row=2, column=col, value=f"=Revenue_By_Channel!{layout.cell(col, 7)}")
        income_sheet.cell(row=3, column=col, value=f"=OPEX!{layout.cell(col, 11)}")
        income_sheet.cell(row=4, column=col, value=f"=Income_Statement!{layout.cell(col, 2)}-Income_Statement!{layout.cell(col, 3)}")
        income_sheet.cell(row=5, column=col, value=f"=Assumptions!{_series_cell(label_depreciation, year)}")
        income_sheet.cell(row=6, column=col, value=f"=Assumptions!{_series_cell(label_grant_release, year)}")
        income_sheet.cell(row=7, column=col, value=f"=Income_Statement!{layout.cell(col, 4)}-Income_Statement!{layout.cell(col, 5)}+Income_Statement!{layout.cell(col, 6)}")
        income_sheet.cell(row=8, column=col, value=f"=Assumptions!{_series_cell(label_interest, year)}")
        income_sheet.cell(row=9, column=col, value=f"=Income_Statement!{layout.cell(col, 7)}-Income_Statement!{layout.cell(col, 8)}")
        income_sheet.cell(row=10, column=col, value=f"=Income_Statement!{layout.cell(col, 9)}*Assumptions!{_param_cell(label_tax_rate)}")
        income_sheet.cell(row=11, column=col, value=f"=Assumptions!{_series_cell(label_tax_credit, year)}")
        income_sheet.cell(row=12, column=col, value=f"=Income_Statement!{layout.cell(col, 10)}-Income_Statement!{layout.cell(col, 11)}")
        income_sheet.cell(row=13, column=col, value=f"=Income_Statement!{layout.cell(col, 9)}-Income_Statement!{layout.cell(col, 12)}")
    _style_table_header(income_sheet, 1, 1 + len(years))
    _style_table_body(income_sheet, 2, 13, 1 + len(years))
    _auto_fit_columns(income_sheet, 1 + len(years))

    balance_sheet = wb.create_sheet("Balance_Sheet")
    _write_year_header(balance_sheet, years, layout, "Line Item")
    balance_labels = [
        _label("Cash", "€"),
        _label("Trade Receivables", "€"),
        _label("Grant Receivable", "€"),
        _label("Tax Credit Receivable", "€"),
        _label("Total Current Assets", "€"),
        _label("Fixed Assets Gross", "€"),
        _label("Accumulated Depreciation", "€"),
        _label("Fixed Assets Net", "€"),
        _label("Total Assets", "€"),
        _label("Trade Payables", "€"),
        _label("Taxes Payable", "€"),
        _label("Total Current Liabilities", "€"),
        _label("Debt", "€"),
        _label("Deferred Income", "€"),
        _label("Total Non-Current Liabilities", "€"),
        _label("Total Liabilities", "€"),
        _label("Share Capital", "€"),
        _label("Retained Earnings", "€"),
        _label("Current Year Profit", "€"),
        _label("Total Equity", "€"),
        _label("Total Liabilities & Equity", "€"),
        _label("Balance Check", "€"),
    ]
    for idx, label in enumerate(balance_labels, start=2):
        balance_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        balance_sheet.cell(row=2, column=col, value=f"=Assumptions!{_series_cell(label_cash, year)}")
        balance_sheet.cell(
            row=3,
            column=col,
            value=(
                f"=Revenue_By_Channel!{layout.cell(col, 2)}*Assumptions!{_param_cell(label_gate_dso)}/365"
                f"+Revenue_By_Channel!{layout.cell(col, 3)}*Assumptions!{_param_cell(label_tariff_dso)}/365"
                f"+Revenue_By_Channel!{layout.cell(col, 4)}*Assumptions!{_param_cell(label_go_dso)}/365"
                f"+Revenue_By_Channel!{layout.cell(col, 5)}*Assumptions!{_param_cell(label_co2_dso)}/365"
                f"+Revenue_By_Channel!{layout.cell(col, 6)}*Assumptions!{_param_cell(label_compost_dso)}/365"
            ),
        )
        balance_sheet.cell(row=4, column=col, value=f"=Assumptions!{_series_cell(label_grant_receivable, year)}")
        balance_sheet.cell(row=5, column=col, value=f"=Assumptions!{_series_cell(label_tax_credit_receivable, year)}")
        balance_sheet.cell(row=6, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 2)}:{layout.cell(col, 5)})")
        balance_sheet.cell(row=7, column=col, value=f"=Assumptions!{_series_cell(label_fixed_assets_gross, year)}")
        balance_sheet.cell(row=8, column=col, value=f"=Assumptions!{_series_cell(label_accum_depr, year)}")
        balance_sheet.cell(row=9, column=col, value=f"=Balance_Sheet!{layout.cell(col, 7)}-Balance_Sheet!{layout.cell(col, 8)}")
        balance_sheet.cell(row=10, column=col, value=f"=Balance_Sheet!{layout.cell(col, 6)}+Balance_Sheet!{layout.cell(col, 9)}")
        balance_sheet.cell(
            row=11,
            column=col,
            value=(
                f"=OPEX!{layout.cell(col, 2)}*Assumptions!{_param_cell(label_feedstock_dpo)}/365"
                f"+OPEX!{layout.cell(col, 3)}*Assumptions!{_param_cell(label_utilities_dpo)}/365"
                f"+OPEX!{layout.cell(col, 4)}*Assumptions!{_param_cell(label_chemicals_dpo)}/365"
                f"+OPEX!{layout.cell(col, 5)}*Assumptions!{_param_cell(label_maintenance_dpo)}/365"
                f"+OPEX!{layout.cell(col, 6)}*Assumptions!{_param_cell(label_personnel_dpo)}/365"
                f"+OPEX!{layout.cell(col, 7)}*Assumptions!{_param_cell(label_insurance_dpo)}/365"
                f"+OPEX!{layout.cell(col, 8)}*Assumptions!{_param_cell(label_overheads_dpo)}/365"
                f"+OPEX!{layout.cell(col, 9)}*Assumptions!{_param_cell(label_digestate_dpo)}/365"
                f"+OPEX!{layout.cell(col, 10)}*Assumptions!{_param_cell(label_other_dpo)}/365"
            ),
        )
        balance_sheet.cell(row=12, column=col, value=f"=Assumptions!{_series_cell(label_taxes_payable, year)}")
        balance_sheet.cell(row=13, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 11)}:{layout.cell(col, 12)})")
        balance_sheet.cell(row=14, column=col, value=f"=Assumptions!{_series_cell(label_debt, year)}")
        balance_sheet.cell(row=15, column=col, value=f"=Assumptions!{_series_cell(label_deferred_income, year)}")
        balance_sheet.cell(row=16, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 14)}:{layout.cell(col, 15)})")
        balance_sheet.cell(row=17, column=col, value=f"=Balance_Sheet!{layout.cell(col, 13)}+Balance_Sheet!{layout.cell(col, 16)}")
        balance_sheet.cell(row=18, column=col, value=f"=Assumptions!{_series_cell(label_share_capital, year)}")
        balance_sheet.cell(row=19, column=col, value=f"=Assumptions!{_series_cell(label_retained_earnings, year)}")
        balance_sheet.cell(row=20, column=col, value=f"=Assumptions!{_series_cell(label_current_profit, year)}")
        balance_sheet.cell(row=21, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 18)}:{layout.cell(col, 20)})")
        balance_sheet.cell(row=22, column=col, value=f"=Balance_Sheet!{layout.cell(col, 17)}+Balance_Sheet!{layout.cell(col, 21)}")
        balance_sheet.cell(row=23, column=col, value=f"=Balance_Sheet!{layout.cell(col, 10)}-Balance_Sheet!{layout.cell(col, 22)}")
    _style_table_header(balance_sheet, 1, 1 + len(years))
    _style_table_body(balance_sheet, 2, 23, 1 + len(years))
    _auto_fit_columns(balance_sheet, 1 + len(years))

    reclass_sheet = wb.create_sheet("Balance_Sheet_Reclass")
    _write_year_header(reclass_sheet, years, layout, "Line Item")
    reclass_labels = [
        _label("Fixed Assets (Net)", "€"),
        _label("NWC", "€"),
        _label("CIN", "€"),
        _label("Debt", "€"),
        _label("Cash", "€"),
        _label("NFP", "€"),
        _label("Equity", "€"),
        _label("CIN - (Equity + NFP)", "€"),
    ]
    for idx, label in enumerate(reclass_labels, start=2):
        reclass_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        reclass_sheet.cell(row=2, column=col, value=f"=Balance_Sheet!{layout.cell(col, 9)}")
        reclass_sheet.cell(
            row=3,
            column=col,
            value=f"=Balance_Sheet!{layout.cell(col, 3)}-Balance_Sheet!{layout.cell(col, 11)}",
        )
        reclass_sheet.cell(row=4, column=col, value=f"=Balance_Sheet_Reclass!{layout.cell(col, 2)}+Balance_Sheet_Reclass!{layout.cell(col, 3)}")
        reclass_sheet.cell(row=5, column=col, value=f"=Balance_Sheet!{layout.cell(col, 14)}")
        reclass_sheet.cell(row=6, column=col, value=f"=Balance_Sheet!{layout.cell(col, 2)}")
        reclass_sheet.cell(row=7, column=col, value=f"=Balance_Sheet_Reclass!{layout.cell(col, 5)}-Balance_Sheet_Reclass!{layout.cell(col, 6)}")
        reclass_sheet.cell(row=8, column=col, value=f"=Balance_Sheet!{layout.cell(col, 21)}")
        reclass_sheet.cell(
            row=9,
            column=col,
            value=f"=Balance_Sheet_Reclass!{layout.cell(col, 4)}-(Balance_Sheet_Reclass!{layout.cell(col, 7)}+Balance_Sheet_Reclass!{layout.cell(col, 8)})",
        )
    _style_table_header(reclass_sheet, 1, 1 + len(years))
    _style_table_body(reclass_sheet, 2, 9, 1 + len(years))
    _auto_fit_columns(reclass_sheet, 1 + len(years))

    cash_flow = wb.create_sheet("Cash_Flow")
    _write_year_header(cash_flow, years, layout, "Line Item")
    cf_labels = [
        _label("Net Income", "€"),
        _label("Interest", "€"),
        _label("Depreciation", "€"),
        _label("Grant Release", "€"),
        _label("Change in NWC", "€"),
        _label("CFO", "€"),
        _label("Capex", "€"),
        _label("Grant Cash", "€"),
        _label("CFI", "€"),
        _label("Debt Draw", "€"),
        _label("Debt Repay", "€"),
        _label("Interest Paid", "€"),
        _label("Equity Contribution", "€"),
        _label("CFF", "€"),
        _label("Net CF", "€"),
        _label("Opening Cash", "€"),
        _label("Closing Cash", "€"),
    ]
    for idx, label in enumerate(cf_labels, start=2):
        cash_flow.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        cash_flow.cell(row=2, column=col, value=f"=Income_Statement!{layout.cell(col, 13)}")
        cash_flow.cell(row=3, column=col, value=f"=Income_Statement!{layout.cell(col, 8)}")
        cash_flow.cell(row=4, column=col, value=f"=Assumptions!{_series_cell(label_depreciation, year)}")
        cash_flow.cell(row=5, column=col, value=f"=Assumptions!{_series_cell(label_grant_release, year)}")
        cash_flow.cell(row=6, column=col, value=f"=Assumptions!{_series_cell(label_delta_nwc, year)}")
        cash_flow.cell(row=7, column=col, value=f"=SUM(Cash_Flow!{layout.cell(col, 2)}:{layout.cell(col, 4)})-Cash_Flow!{layout.cell(col, 5)}+Cash_Flow!{layout.cell(col, 6)}")
        cash_flow.cell(row=8, column=col, value=f"=Assumptions!{_series_cell(label_capex, year)}")
        cash_flow.cell(row=9, column=col, value=f"=Assumptions!{_series_cell(label_grant_cash, year)}")
        cash_flow.cell(row=10, column=col, value=f"=Cash_Flow!{layout.cell(col, 8)}+Cash_Flow!{layout.cell(col, 9)}")
        cash_flow.cell(row=11, column=col, value=f"=Assumptions!{_series_cell(label_debt_draw, year)}")
        cash_flow.cell(row=12, column=col, value=f"=Assumptions!{_series_cell(label_debt_repay, year)}")
        cash_flow.cell(row=13, column=col, value=f"=Assumptions!{_series_cell(label_interest_paid, year)}")
        cash_flow.cell(row=14, column=col, value=f"=Assumptions!{_series_cell(label_equity_contrib, year)}")
        cash_flow.cell(row=15, column=col, value=f"=SUM(Cash_Flow!{layout.cell(col, 11)}:{layout.cell(col, 14)})")
        cash_flow.cell(row=16, column=col, value=f"=Cash_Flow!{layout.cell(col, 7)}+Cash_Flow!{layout.cell(col, 10)}+Cash_Flow!{layout.cell(col, 15)}")
        cash_flow.cell(row=17, column=col, value=f"=Assumptions!{_series_cell(label_opening_cash, year)}")
        cash_flow.cell(row=18, column=col, value=f"=Cash_Flow!{layout.cell(col, 17)}+Cash_Flow!{layout.cell(col, 16)}")
    _style_table_header(cash_flow, 1, 1 + len(years))
    _style_table_body(cash_flow, 2, 18, 1 + len(years))
    _auto_fit_columns(cash_flow, 1 + len(years))

    fcff_sheet = wb.create_sheet("FCFF")
    _write_year_header(fcff_sheet, operating_years, layout, "Line Item")
    fcff_labels = [
        _label("EBIT", "€"),
        _label("D&A", "€"),
        _label("ΔNWC", "€"),
        _label("Capex", "€"),
        _label("FCFF", "€"),
        _label("Net Borrowing", "€"),
        _label("FCFE", "€"),
    ]
    for idx, label in enumerate(fcff_labels, start=2):
        fcff_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(operating_years).items():
        fcff_sheet.cell(row=2, column=col, value=f"=Income_Statement!{layout.cell(layout.year_to_col(years)[year], 7)}")
        fcff_sheet.cell(row=3, column=col, value=f"=Assumptions!{_series_cell(label_depreciation, year)}")
        fcff_sheet.cell(row=4, column=col, value=f"=Assumptions!{_series_cell(label_delta_nwc, year)}")
        fcff_sheet.cell(row=5, column=col, value=f"=Assumptions!{_series_cell(label_capex, year)}")
        fcff_sheet.cell(row=6, column=col, value=f"=FCFF!{layout.cell(col, 2)}*(1-Assumptions!{_param_cell(label_tax_rate)})+FCFF!{layout.cell(col, 3)}-FCFF!{layout.cell(col, 4)}-FCFF!{layout.cell(col, 5)}")
        fcff_sheet.cell(row=7, column=col, value=f"=Cash_Flow!{layout.cell(layout.year_to_col(years)[year], 11)}+Cash_Flow!{layout.cell(layout.year_to_col(years)[year], 12)}")
        fcff_sheet.cell(row=8, column=col, value=f"=FCFF!{layout.cell(col, 6)}-Income_Statement!{layout.cell(layout.year_to_col(years)[year], 8)}+Income_Statement!{layout.cell(layout.year_to_col(years)[year], 8)}*Assumptions!{_param_cell(label_tax_rate)}+FCFF!{layout.cell(col, 7)}")
    _style_table_header(fcff_sheet, 1, 1 + len(operating_years))
    _style_table_body(fcff_sheet, 2, 8, 1 + len(operating_years))
    _auto_fit_columns(fcff_sheet, 1 + len(operating_years))

    discounting = wb.create_sheet("Discounting")
    _write_year_header(discounting, operating_years, layout, "Line Item")
    disc_labels = [
        _label("Period", "year"),
        _label("WACC", "%"),
        _label("Ke", "%"),
        _label("DF (WACC)", "x"),
        _label("DF (Ke)", "x"),
        _label("FCFF", "€"),
        _label("PV(FCFF)", "€"),
        _label("FCFE", "€"),
        _label("PV(FCFE)", "€"),
    ]
    for idx, label in enumerate(disc_labels, start=2):
        discounting.cell(row=idx, column=1, value=label)
    for idx, (year, col) in enumerate(layout.year_to_col(operating_years).items(), start=1):
        discounting.cell(row=2, column=col, value=idx)
        discounting.cell(row=3, column=col, value=valuation.wacc.get(year, 0.0))
        discounting.cell(row=4, column=col, value=f"=Assumptions!{_param_cell(label_ke)}")
        discounting.cell(row=5, column=col, value=f"=1/(1+Discounting!{layout.cell(col, 3)})^Discounting!{layout.cell(col, 2)}")
        discounting.cell(row=6, column=col, value=f"=1/(1+Discounting!{layout.cell(col, 4)})^Discounting!{layout.cell(col, 2)}")
        discounting.cell(row=7, column=col, value=f"=FCFF!{layout.cell(col, 6)}")
        discounting.cell(row=8, column=col, value=f"=Discounting!{layout.cell(col, 7)}*Discounting!{layout.cell(col, 5)}")
        discounting.cell(row=9, column=col, value=f"=FCFF!{layout.cell(col, 8)}")
        discounting.cell(row=10, column=col, value=f"=Discounting!{layout.cell(col, 9)}*Discounting!{layout.cell(col, 6)}")
    _style_table_header(discounting, 1, 1 + len(operating_years))
    _style_table_body(discounting, 2, 10, 1 + len(operating_years))
    _auto_fit_columns(discounting, 1 + len(operating_years))


    opex_sheet = wb.create_sheet("OPEX")
    _write_year_header(opex_sheet, years, layout, "Line Item")
    opex_labels = [
        ("Feedstock", "OPEX Feedstock"),
        ("Utilities", "OPEX Utilities"),
        ("Chemicals", "OPEX Chemicals"),
        ("Maintenance", "OPEX Maintenance"),
        ("Personnel", "OPEX Personnel"),
        ("Insurance", "OPEX Insurance"),
        ("Overheads", "OPEX Overheads"),
        ("Digestate", "OPEX Digestate"),
        ("Other", "OPEX Other"),
        ("Total", None),
    ]
    for idx, (label, _) in enumerate(opex_labels, start=2):
        opex_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        for idx, (_, series_label) in enumerate(opex_labels, start=2):
            if series_label:
                opex_sheet.cell(row=idx, column=col, value=f"=Assumptions!{_series_cell(series_label, year)}")
        opex_sheet.cell(
            row=11,
            column=col,
            value=f"=SUM(OPEX!{layout.cell(col, 2)}:{layout.cell(col, 10)})",
        )
    _style_table_header(opex_sheet, 1, 1 + len(years))
    _style_table_body(opex_sheet, 2, 11, 1 + len(years))
    _auto_fit_columns(opex_sheet, 1 + len(years))

    income_sheet = wb.create_sheet("Income_Statement")
    _write_year_header(income_sheet, years, layout, "Line Item")
    income_labels = [
        "Revenue",
        "OPEX",
        "EBITDA",
        "D&A",
        "Grant Release",
        "EBIT",
        "Interest",
        "EBT",
        "Taxes Before Credit",
        "Tax Credit Used",
        "Taxes Paid",
        "Net Income",
    ]
    for idx, label in enumerate(income_labels, start=2):
        income_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        income_sheet.cell(row=2, column=col, value=f"=Revenue_By_Channel!{layout.cell(col, 7)}")
        income_sheet.cell(row=3, column=col, value=f"=OPEX!{layout.cell(col, 11)}")
        income_sheet.cell(row=4, column=col, value=f"=Income_Statement!{layout.cell(col, 2)}-Income_Statement!{layout.cell(col, 3)}")
        income_sheet.cell(row=5, column=col, value=f"=Assumptions!{_series_cell('Depreciation', year)}")
        income_sheet.cell(row=6, column=col, value=f"=Assumptions!{_series_cell('Grant Release', year)}")
        income_sheet.cell(row=7, column=col, value=f"=Income_Statement!{layout.cell(col, 4)}-Income_Statement!{layout.cell(col, 5)}+Income_Statement!{layout.cell(col, 6)}")
        income_sheet.cell(row=8, column=col, value=f"=Assumptions!{_series_cell('Interest Expense', year)}")
        income_sheet.cell(row=9, column=col, value=f"=Income_Statement!{layout.cell(col, 7)}-Income_Statement!{layout.cell(col, 8)}")
        income_sheet.cell(row=10, column=col, value=f"=Income_Statement!{layout.cell(col, 9)}*Assumptions!{_param_cell('Tax Rate')}")
        income_sheet.cell(row=11, column=col, value=f"=Assumptions!{_series_cell('Tax Credit Used', year)}")
        income_sheet.cell(row=12, column=col, value=f"=Income_Statement!{layout.cell(col, 10)}-Income_Statement!{layout.cell(col, 11)}")
        income_sheet.cell(row=13, column=col, value=f"=Income_Statement!{layout.cell(col, 9)}-Income_Statement!{layout.cell(col, 12)}")
    _style_table_header(income_sheet, 1, 1 + len(years))
    _style_table_body(income_sheet, 2, 13, 1 + len(years))
    _auto_fit_columns(income_sheet, 1 + len(years))

    balance_sheet = wb.create_sheet("Balance_Sheet")
    _write_year_header(balance_sheet, years, layout, "Line Item")
    balance_labels = [
        "Cash",
        "Trade Receivables",
        "Grant Receivable",
        "Tax Credit Receivable",
        "Total Current Assets",
        "Fixed Assets Gross",
        "Accumulated Depreciation",
        "Fixed Assets Net",
        "Total Assets",
        "Trade Payables",
        "Taxes Payable",
        "Total Current Liabilities",
        "Debt",
        "Deferred Income",
        "Total Non-Current Liabilities",
        "Total Liabilities",
        "Share Capital",
        "Retained Earnings",
        "Current Year Profit",
        "Total Equity",
        "Total Liabilities & Equity",
        "Balance Check",
    ]
    for idx, label in enumerate(balance_labels, start=2):
        balance_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        balance_sheet.cell(row=2, column=col, value=f"=Assumptions!{_series_cell('Cash', year)}")
        balance_sheet.cell(row=3, column=col, value=f"=Assumptions!{_series_cell('Trade Receivables', year)}")
        balance_sheet.cell(row=4, column=col, value=f"=Assumptions!{_series_cell('Grant Receivable', year)}")
        balance_sheet.cell(row=5, column=col, value=f"=Assumptions!{_series_cell('Tax Credit Receivable', year)}")
        balance_sheet.cell(row=6, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 2)}:{layout.cell(col, 5)})")
        balance_sheet.cell(row=7, column=col, value=f"=Assumptions!{_series_cell('Fixed Assets Gross', year)}")
        balance_sheet.cell(row=8, column=col, value=f"=Assumptions!{_series_cell('Accumulated Depreciation', year)}")
        balance_sheet.cell(row=9, column=col, value=f"=Balance_Sheet!{layout.cell(col, 7)}-Balance_Sheet!{layout.cell(col, 8)}")
        balance_sheet.cell(row=10, column=col, value=f"=Balance_Sheet!{layout.cell(col, 6)}+Balance_Sheet!{layout.cell(col, 9)}")
        balance_sheet.cell(row=11, column=col, value=f"=Assumptions!{_series_cell('Trade Payables', year)}")
        balance_sheet.cell(row=12, column=col, value=f"=Assumptions!{_series_cell('Taxes Payable', year)}")
        balance_sheet.cell(row=13, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 11)}:{layout.cell(col, 12)})")
        balance_sheet.cell(row=14, column=col, value=f"=Assumptions!{_series_cell('Debt', year)}")
        balance_sheet.cell(row=15, column=col, value=f"=Assumptions!{_series_cell('Deferred Income', year)}")
        balance_sheet.cell(row=16, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 14)}:{layout.cell(col, 15)})")
        balance_sheet.cell(row=17, column=col, value=f"=Balance_Sheet!{layout.cell(col, 13)}+Balance_Sheet!{layout.cell(col, 16)}")
        balance_sheet.cell(row=18, column=col, value=f"=Assumptions!{_series_cell('Share Capital', year)}")
        balance_sheet.cell(row=19, column=col, value=f"=Assumptions!{_series_cell('Retained Earnings', year)}")
        balance_sheet.cell(row=20, column=col, value=f"=Assumptions!{_series_cell('Current Year Profit', year)}")
        balance_sheet.cell(row=21, column=col, value=f"=SUM(Balance_Sheet!{layout.cell(col, 18)}:{layout.cell(col, 20)})")
        balance_sheet.cell(row=22, column=col, value=f"=Balance_Sheet!{layout.cell(col, 17)}+Balance_Sheet!{layout.cell(col, 21)}")
        balance_sheet.cell(row=23, column=col, value=f"=Balance_Sheet!{layout.cell(col, 10)}-Balance_Sheet!{layout.cell(col, 22)}")
    _style_table_header(balance_sheet, 1, 1 + len(years))
    _style_table_body(balance_sheet, 2, 23, 1 + len(years))
    _auto_fit_columns(balance_sheet, 1 + len(years))

    cash_flow = wb.create_sheet("Cash_Flow")
    _write_year_header(cash_flow, years, layout, "Line Item")
    cf_labels = [
        "Net Income",
        "Interest",
        "Depreciation",
        "Grant Release",
        "Change in NWC",
        "CFO",
        "Capex",
        "Grant Cash",
        "CFI",
        "Debt Draw",
        "Debt Repay",
        "Interest Paid",
        "Equity Contribution",
        "CFF",
        "Net CF",
        "Opening Cash",
        "Closing Cash",
    ]
    for idx, label in enumerate(cf_labels, start=2):
        cash_flow.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(years).items():
        cash_flow.cell(row=2, column=col, value=f"=Income_Statement!{layout.cell(col, 13)}")
        cash_flow.cell(row=3, column=col, value=f"=Income_Statement!{layout.cell(col, 8)}")
        cash_flow.cell(row=4, column=col, value=f"=Assumptions!{_series_cell('Depreciation', year)}")
        cash_flow.cell(row=5, column=col, value=f"=Assumptions!{_series_cell('Grant Release', year)}")
        cash_flow.cell(row=6, column=col, value=f"=Assumptions!{_series_cell('Delta NWC', year)}")
        cash_flow.cell(row=7, column=col, value=f"=SUM(Cash_Flow!{layout.cell(col, 2)}:{layout.cell(col, 4)})-Cash_Flow!{layout.cell(col, 5)}+Cash_Flow!{layout.cell(col, 6)}")
        cash_flow.cell(row=8, column=col, value=f"=Assumptions!{_series_cell('Capex', year)}")
        cash_flow.cell(row=9, column=col, value=f"=Assumptions!{_series_cell('Grant Cash', year)}")
        cash_flow.cell(row=10, column=col, value=f"=Cash_Flow!{layout.cell(col, 8)}+Cash_Flow!{layout.cell(col, 9)}")
        cash_flow.cell(row=11, column=col, value=f"=Assumptions!{_series_cell('Debt Draw', year)}")
        cash_flow.cell(row=12, column=col, value=f"=Assumptions!{_series_cell('Debt Repay', year)}")
        cash_flow.cell(row=13, column=col, value=f"=Assumptions!{_series_cell('Interest Paid', year)}")
        cash_flow.cell(row=14, column=col, value=f"=Assumptions!{_series_cell('Equity Contribution', year)}")
        cash_flow.cell(row=15, column=col, value=f"=SUM(Cash_Flow!{layout.cell(col, 11)}:{layout.cell(col, 14)})")
        cash_flow.cell(row=16, column=col, value=f"=Cash_Flow!{layout.cell(col, 7)}+Cash_Flow!{layout.cell(col, 10)}+Cash_Flow!{layout.cell(col, 15)}")
        cash_flow.cell(row=17, column=col, value=f"=Assumptions!{_series_cell('Opening Cash', year)}")
        cash_flow.cell(row=18, column=col, value=f"=Cash_Flow!{layout.cell(col, 17)}+Cash_Flow!{layout.cell(col, 16)}")
    _style_table_header(cash_flow, 1, 1 + len(years))
    _style_table_body(cash_flow, 2, 18, 1 + len(years))
    _auto_fit_columns(cash_flow, 1 + len(years))

    fcff_sheet = wb.create_sheet("FCFF")
    _write_year_header(fcff_sheet, operating_years, layout, "Line Item")
    fcff_labels = ["EBIT", "D&A", "ΔNWC", "Capex", "FCFF", "Net Borrowing", "FCFE"]
    for idx, label in enumerate(fcff_labels, start=2):
        fcff_sheet.cell(row=idx, column=1, value=label)
    for year, col in layout.year_to_col(operating_years).items():
        fcff_sheet.cell(row=2, column=col, value=f"=Income_Statement!{layout.cell(layout.year_to_col(years)[year], 7)}")
        fcff_sheet.cell(row=3, column=col, value=f"=Assumptions!{_series_cell('Depreciation', year)}")
        fcff_sheet.cell(row=4, column=col, value=f"=Assumptions!{_series_cell('Delta NWC', year)}")
        fcff_sheet.cell(row=5, column=col, value=f"=Assumptions!{_series_cell('Capex', year)}")
        fcff_sheet.cell(row=6, column=col, value=f"=FCFF!{layout.cell(col, 2)}*(1-Assumptions!{_param_cell('Tax Rate')})+FCFF!{layout.cell(col, 3)}-FCFF!{layout.cell(col, 4)}-FCFF!{layout.cell(col, 5)}")
        fcff_sheet.cell(row=7, column=col, value=f"=Cash_Flow!{layout.cell(layout.year_to_col(years)[year], 11)}+Cash_Flow!{layout.cell(layout.year_to_col(years)[year], 12)}")
        fcff_sheet.cell(row=8, column=col, value=f"=FCFF!{layout.cell(col, 6)}-Income_Statement!{layout.cell(layout.year_to_col(years)[year], 8)}+Income_Statement!{layout.cell(layout.year_to_col(years)[year], 8)}*Assumptions!{_param_cell('Tax Rate')}+FCFF!{layout.cell(col, 7)}")
    _style_table_header(fcff_sheet, 1, 1 + len(operating_years))
    _style_table_body(fcff_sheet, 2, 8, 1 + len(operating_years))
    _auto_fit_columns(fcff_sheet, 1 + len(operating_years))

    discounting = wb.create_sheet("Discounting")
    _write_year_header(discounting, operating_years, layout, "Line Item")
    disc_labels = [
        "Period",
        "WACC",
        "Ke",
        "DF (WACC)",
        "DF (Ke)",
        "FCFF",
        "PV(FCFF)",
        "FCFE",
        "PV(FCFE)",
    ]
    for idx, label in enumerate(disc_labels, start=2):
        discounting.cell(row=idx, column=1, value=label)
    for idx, (year, col) in enumerate(layout.year_to_col(operating_years).items(), start=1):
        discounting.cell(row=2, column=col, value=idx)
        discounting.cell(row=3, column=col, value=valuation.wacc.get(year, 0.0))
        discounting.cell(row=4, column=col, value=f"=Assumptions!{_param_cell('Cost of Equity (Ke)')}")
        discounting.cell(row=5, column=col, value=f"=1/(1+Discounting!{layout.cell(col, 3)})^Discounting!{layout.cell(col, 2)}")
        discounting.cell(row=6, column=col, value=f"=1/(1+Discounting!{layout.cell(col, 4)})^Discounting!{layout.cell(col, 2)}")
        discounting.cell(row=7, column=col, value=f"=FCFF!{layout.cell(col, 6)}")
        discounting.cell(row=8, column=col, value=f"=Discounting!{layout.cell(col, 7)}*Discounting!{layout.cell(col, 5)}")
        discounting.cell(row=9, column=col, value=f"=FCFF!{layout.cell(col, 8)}")
        discounting.cell(row=10, column=col, value=f"=Discounting!{layout.cell(col, 9)}*Discounting!{layout.cell(col, 6)}")
    _style_table_header(discounting, 1, 1 + len(operating_years))
    _style_table_body(discounting, 2, 10, 1 + len(operating_years))
    _auto_fit_columns(discounting, 1 + len(operating_years))

    valuation_sheet = wb.create_sheet("Valuation_Summary")
    valuation_sheet["A1"] = "Metric"
    valuation_sheet["B1"] = "Value"
    metrics = [
        _label("Sum PV(FCFF)", "€"),
        _label("Terminal Value (FCFF)", "€"),
        _label("PV Terminal Value (FCFF)", "€"),
        _label("Enterprise Value", "€"),
        _label("Debt at Base", "€"),
        _label("Cash at Base", "€"),
        _label("Net Debt", "€"),
        _label("Equity Value (FCFF/WACC)", "€"),
        _label("Sum PV(FCFE)", "€"),
        _label("PV Terminal Value (FCFE)", "€"),
        _label("Equity Value (Direct)", "€"),
        _label("Reconciliation Difference", "€"),
        "Sum PV(FCFF)",
        "Terminal Value (FCFF)",
        "PV Terminal Value (FCFF)",
        "Enterprise Value",
        "Debt at Base",
        "Cash at Base",
        "Net Debt",
        "Equity Value (FCFF/WACC)",
        "Sum PV(FCFE)",
        "PV Terminal Value (FCFE)",
        "Equity Value (Direct)",
        "Reconciliation Difference",
    ]
    for idx, label in enumerate(metrics, start=2):
        valuation_sheet.cell(row=idx, column=1, value=label)
    last_col = layout.year_to_col(operating_years)[operating_years[-1]]
    if case.terminal_value.method == "perpetuity":
        tv_formula = (
            f"=FCFF!{layout.cell(last_col, 6)}*(1+Assumptions!{_param_cell(label_g)})"
            f"/(Discounting!{layout.cell(last_col, 3)}-Assumptions!{_param_cell(label_g)})"
            f"=FCFF!{layout.cell(last_col, 6)}*(1+Assumptions!{_param_cell('Terminal Growth Rate (g)')})"
            f"/(Discounting!{layout.cell(last_col, 3)}-Assumptions!{_param_cell('Terminal Growth Rate (g)')})"
        )
    else:
        tv_formula = (
            f"=Income_Statement!{layout.cell(layout.year_to_col(years)[operating_years[-1]], 4)}"
            f"*Assumptions!{_param_cell(label_exit_multiple)}"
            f"*Assumptions!{_param_cell('Exit Multiple')}"
        )
    valuation_sheet.cell(
        row=2,
        column=2,
        value=f"=SUM(Discounting!{layout.cell(layout.start_col, 8)}:{layout.cell(last_col, 8)})",
    )
    valuation_sheet.cell(row=3, column=2, value=tv_formula)
    valuation_sheet.cell(row=4, column=2, value=f"=Valuation_Summary!B3*Discounting!{layout.cell(last_col, 5)}")
    valuation_sheet.cell(row=5, column=2, value="=Valuation_Summary!B2+Valuation_Summary!B4")
    valuation_sheet.cell(row=6, column=2, value=f"=Assumptions!{_param_cell(label_debt_base)}")
    valuation_sheet.cell(row=7, column=2, value=f"=Assumptions!{_param_cell(label_cash_base)}")
    valuation_sheet.cell(row=6, column=2, value=f"=Assumptions!{_param_cell('Debt at Base')}")
    valuation_sheet.cell(row=7, column=2, value=f"=Assumptions!{_param_cell('Cash at Base')}")
    valuation_sheet.cell(row=8, column=2, value="=Valuation_Summary!B6-Valuation_Summary!B7")
    valuation_sheet.cell(row=9, column=2, value="=Valuation_Summary!B5-Valuation_Summary!B8")
    valuation_sheet.cell(
        row=10,
        column=2,
        value=f"=SUM(Discounting!{layout.cell(layout.start_col, 10)}:{layout.cell(last_col, 10)})",
    )
    valuation_sheet.cell(row=11, column=2, value=f"=Valuation_Summary!B3*Discounting!{layout.cell(last_col, 6)}")
    valuation_sheet.cell(row=12, column=2, value="=Valuation_Summary!B10+Valuation_Summary!B11")
    valuation_sheet.cell(row=13, column=2, value="=Valuation_Summary!B8-Valuation_Summary!B12")
    _style_table_header(valuation_sheet, 1, 2)
    _style_table_body(valuation_sheet, 2, 13, 2)
    _auto_fit_columns(valuation_sheet, 2)

    audit_checks = wb.create_sheet("Audit_Checks")
    _write_year_header(audit_checks, years, layout, "Check")
    audit_checks.cell(row=2, column=1, value=_label("BS Check", "€"))
    audit_checks.cell(row=3, column=1, value=_label("CIN Identity", "€"))
    audit_checks.cell(row=4, column=1, value=_label("Cash Check", "€"))
    audit_checks.cell(row=5, column=1, value=_label("FCFF Identity", "€"))
    audit_checks.cell(row=6, column=1, value=_label("PV Roll-up", "€"))
    for year, col in layout.year_to_col(years).items():
        audit_checks.cell(row=2, column=col, value=f"=Balance_Sheet!{layout.cell(col, 23)}")
        audit_checks.cell(row=3, column=col, value=f"=Balance_Sheet_Reclass!{layout.cell(col, 9)}")
        if year == years[0]:
            audit_checks.cell(
                row=6,
                column=col,
                value="=Valuation_Summary!B5-(Valuation_Summary!B2+Valuation_Summary!B4)",
            )
        else:
            audit_checks.cell(row=6, column=col, value="")
        if year == years[0]:
            audit_checks.cell(
                row=4,
                column=col,
                value=f"=Cash_Flow!{layout.cell(col, 17)}+Cash_Flow!{layout.cell(col, 16)}-Cash_Flow!{layout.cell(col, 18)}",
            )
        else:
            prev_col = layout.year_to_col(years)[years[years.index(year) - 1]]
            audit_checks.cell(
                row=4,
                column=col,
                value=f"=Cash_Flow!{layout.cell(prev_col, 18)}+Cash_Flow!{layout.cell(col, 16)}-Cash_Flow!{layout.cell(col, 18)}",
            )
        audit_checks.cell(
            row=5,
            column=col,
            value=(
                f"=FCFF!{layout.cell(layout.year_to_col(operating_years).get(year, col), 6)}"
                f"-(Income_Statement!{layout.cell(col, 7)}*(1-Assumptions!{_param_cell(label_tax_rate)})"
                f"+Assumptions!{_series_cell(label_depreciation, year)}"
                f"-Assumptions!{_series_cell(label_capex, year)}"
                f"-Assumptions!{_series_cell(label_delta_nwc, year)})"
            )
            if year in operating_years
            else "",
        )
    _style_table_header(audit_checks, 1, 1 + len(years))
    _style_table_body(audit_checks, 2, 6, 1 + len(years))
    _auto_fit_columns(audit_checks, 1 + len(years))

    audit = wb.create_sheet("Audit_Notes")
    _write_audit_notes(
        audit,
        [
            "Assumptions",
            "Production",
            "Revenue_By_Channel",
            "OPEX",
            "Income_Statement",
            "Balance_Sheet",
            "Balance_Sheet_Reclass",
            "Cash_Flow",
            "FCFF",
            "Discounting",
            "Valuation_Summary",
            "Audit_Checks",
        ],
        [
            "Assumptions series rows are inputs copied from model outputs (OPEX categories, depreciation, grant release, interest, capex, financing, balance sheet components).",
        ],
        [
            "Assumptions series rows are inputs copied from model outputs (OPEX categories, depreciation, grant release, interest, NWC, capex, financing, balance sheet components).",
        ],
        [
            "Discount factors use end-of-period convention: DF=1/(1+r)^period.",
            "Terminal value uses case method (perpetuity or exit multiple).",
            "NWC uses payment-delay-driven receivables/payables formulas.",
        ],
    )


def export_xlsx_biometano(
    projections,
    statements,
    valuation,
    path: str | Path,
    *,
    case: BiometanoCase | None = None,
    xlsx_mode: str = "formulas",
) -> None:
    """Export biometano outputs to Excel."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_mode not in {"formulas", "values"}:
        raise ValueError("xlsx_mode must be 'formulas' or 'values'")

    if xlsx_mode == "values":
        tables = format_biometano_tables(projections, statements, valuation)
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, df in tables.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            _style_xlsx_sheet(ws, df)
        wb.save(path)
        return

    if case is None:
        raise ValueError("case is required for formulas mode export")

    wb = Workbook()
    wb.remove(wb.active)
    _write_biometano_formula_workbook(case, projections, statements, valuation, wb)
    wb.save(path)


def export_csv_biometano(projections, statements, valuation, output_dir: str | Path) -> list[Path]:
    """Export biometano outputs to CSV files."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    tables = format_biometano_tables(projections, statements, valuation)
    created_files = []
    
    for table_name, df in tables.items():
        file_path = output_dir / f"{table_name}.csv"
        df.to_csv(file_path, index=False)
        created_files.append(file_path)
    
    return created_files
